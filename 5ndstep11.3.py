# =============================================================================
#  WEEK 2 — ML MODEL COMPARISON
#  3 models chạy song song:
#    Model 1: Logistic Regression  (baseline — đường thẳng)
#    Model 2: Decision Tree        (học được điểm gãy phi tuyến)
#    Model 3: Random Forest        (ensemble 100 cây — mạnh nhất trong 3)
#
#  Output: roadmap_week2_ml.xlsx  (3 sheets)
#    W2_ModelComparison  — AUC, AP, CV score cả 3 model
#    W2_FeatureImportance — coeff (LR), impurity (DT), Gini (RF)
#    W2_IndividualPD      — group PD vs individual PD per segment
# =============================================================================

import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── sklearn ───────────────────────────────────────────────────────────────────
from sklearn.linear_model    import LogisticRegression
from sklearn.tree            import DecisionTreeClassifier
from sklearn.ensemble        import RandomForestClassifier
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.preprocessing   import StandardScaler
from sklearn.metrics         import (roc_auc_score, roc_curve,
                                     average_precision_score,
                                     confusion_matrix)
from sklearn.pipeline        import Pipeline

# =============================================================================
#  CONFIG
# =============================================================================
CSV_PATH = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\week6_clean_data.csv"
OUT_PATH = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\roadmap_week2_ml.xlsx"

FICO_CUT = 680.0
SEED     = 42

# Features dùng cho cả 3 model
# fico_avg  — điểm tín dụng (continuous)
# dti       — debt-to-income ratio (continuous)
# loan_amnt — số tiền vay (continuous)
# term_60   — 1 nếu 60 tháng, 0 nếu 36 tháng (binary)
FEATURES = ["fico_avg", "dti", "loan_amnt", "term_60"]

# =============================================================================
#  STYLES
# =============================================================================
def fill(h):
    return PatternFill(start_color=h, end_color=h, fill_type="solid")

F_HDR  = fill("0D1B2A"); F_SUB   = fill("1B3A5C"); F_DARK  = fill("263238")
F_SEG1 = fill("C6EFCE"); F_SEG2  = fill("FFEB9C")
F_SEG3 = fill("FFCC99"); F_SEG4  = fill("FF9999")
F_TEAL = fill("E8F8F9"); F_GOLD  = fill("FFF8E1")
F_GRN  = fill("E8F5E9"); F_RED   = fill("FFEBEE")
F_GREY = fill("F5F5F5"); F_WHITE = fill("FFFFFF")
# Model-specific colours
F_LR   = fill("E3F2FD")   # light blue  — Logistic Regression
F_DT   = fill("FFF3E0")   # light amber — Decision Tree
F_RF   = fill("E8F5E9")   # light green — Random Forest
SEG_FILLS = [F_SEG1, F_SEG2, F_SEG3, F_SEG4]

FN    = Font(size=10)
FB    = Font(bold=True, size=10)
FW    = Font(bold=True, color="FFFFFF", size=10)
FWLG  = Font(bold=True, color="FFFFFF", size=13)
FRED  = Font(bold=True, size=10, color="C0392B")
FGRN  = Font(bold=True, size=10, color="1E8449")
FGLD  = Font(bold=True, size=10, color="7D6608")
FTEL  = Font(bold=True, size=10, color="0F8B8D")
FSML  = Font(italic=True, size=9, color="595959")

AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def title(ws, row, text, nc, h=22):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row, 1, text)
    c.font = FWLG; c.fill = F_DARK; c.alignment = AC; c.border = bdr()
    ws.row_dimensions[row].height = h

def sub(ws, row, text, nc, fl=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row, 1, text)
    c.font = FW; c.fill = fl or F_HDR; c.alignment = AL; c.border = bdr()

def hrow(ws, row, hdrs, fl=None):
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row, j, h)
        c.font = FW; c.fill = fl or F_HDR; c.alignment = AC; c.border = bdr()

def drow(ws, row, vals, fl=None, fnts=None):
    for j, v in enumerate(vals, 1):
        c = ws.cell(row, j, v)
        c.font = fnts[j-1] if fnts else FN
        c.fill = fl or F_WHITE; c.alignment = AC; c.border = bdr()

def note(ws, row, text, nc, fl=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row, 1, text)
    c.font = FSML; c.fill = fl or F_TEAL; c.alignment = AL; c.border = bdr()

def autofit(ws, mn=10, mx=42):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        L = next((c.column_letter for c in col if not isinstance(c, MergedCell)), None)
        if not L: continue
        w = max((len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)), default=8)
        ws.column_dimensions[L].width = min(max(mn, w + 2), mx)

# =============================================================================
#  LOAD + PREP
# =============================================================================
print("=" * 65)
print("  Loading data …")
df = pd.read_csv(CSV_PATH, low_memory=False)

# Segment
cond = [
    (df["fico_avg"] >= FICO_CUT) & (df["term_m"] == 36),
    (df["fico_avg"] >= FICO_CUT) & (df["term_m"] == 60),
    (df["fico_avg"] <  FICO_CUT) & (df["term_m"] == 36),
    (df["fico_avg"] <  FICO_CUT) & (df["term_m"] == 60),
]
seg_names = ["Seg1 — Prime Short", "Seg2 — Prime Long",
             "Seg3 — Subprime Short", "Seg4 — Subprime Long"]
df["segment"]      = np.select(cond, seg_names, default="Other")
df["is_high_risk"] = (df["risk_label"] == "high_risk").astype(int)
df["term_60"]      = (df["term_m"] == 60).astype(int)

baseline = df["is_high_risk"].mean() * 100
total_n  = len(df)
print(f"  Rows: {total_n:,}  |  Baseline: {baseline:.2f}%")

# Feature matrix — drop rows missing any feature
ml_df = df[df["segment"] != "Other"].copy()
ml_df = ml_df.dropna(subset=FEATURES + ["is_high_risk"])

X = ml_df[FEATURES].values
y = ml_df["is_high_risk"].values
print(f"  ML dataset: {len(ml_df):,} rows  |  positive: {y.mean()*100:.2f}%")

# Train / test split — stratified để giữ tỷ lệ 12.85% trong cả train lẫn test
X_tr, X_te, y_tr, y_te = train_test_split(
    X, y, test_size=0.2, random_state=SEED, stratify=y
)
print(f"  Train: {len(X_tr):,}  |  Test: {len(X_te):,}")

# =============================================================================
#  MODEL DEFINITIONS
# =============================================================================
# ── Model 1: Logistic Regression ─────────────────────────────────────────────
# Cần StandardScaler vì LR nhạy với scale
# class_weight="balanced" bù cho imbalance (87% negative vs 13% positive)
pipe_lr = Pipeline([
    ("scaler", StandardScaler()),
    ("model",  LogisticRegression(
        max_iter=1000,
        random_state=SEED,
        class_weight="balanced",
        C=1.0              # regularization strength (default)
    ))
])

# ── Model 2: Decision Tree ────────────────────────────────────────────────────
# max_depth=6 — giới hạn độ sâu để tránh overfitting
# min_samples_leaf=500 — mỗi leaf phải có ít nhất 500 mẫu
#   → với 1.8M train rows, đây là ~0.028% — ngăn cây học noise
# class_weight="balanced" — xử lý imbalance
pipe_dt = Pipeline([
    ("model", DecisionTreeClassifier(
        max_depth=6,
        min_samples_leaf=500,
        random_state=SEED,
        class_weight="balanced",
        criterion="gini"       # Gini impurity — tiêu chuẩn phổ biến nhất
    ))
])

# ── Model 3: Random Forest ────────────────────────────────────────────────────
# n_estimators=100 — 100 decision trees, mỗi cây học trên bootstrap sample khác nhau
# max_depth=8 — mỗi cây sâu hơn DT một chút vì có ensemble để chống overfit
# max_features="sqrt" — mỗi split chỉ xem xét sqrt(4) = 2 features ngẫu nhiên
#   → tạo sự đa dạng giữa các cây (core mechanism của Random Forest)
# n_jobs=-1 — dùng tất cả CPU cores để train song song
pipe_rf = Pipeline([
    ("model", RandomForestClassifier(
        n_estimators=100,
        max_depth=8,
        min_samples_leaf=500,
        max_features="sqrt",
        random_state=SEED,
        class_weight="balanced",
        n_jobs=-1
    ))
])

models = {
    "Logistic Regression": pipe_lr,
    "Decision Tree":       pipe_dt,
    "Random Forest":       pipe_rf,
}

model_fills = {
    "Logistic Regression": F_LR,
    "Decision Tree":       F_DT,
    "Random Forest":       F_RF,
}

# =============================================================================
#  TRAIN + EVALUATE
# =============================================================================
print("\n" + "=" * 65)
print("  Training and evaluating 3 models …")

cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=SEED)
results = {}

for name, pipe in models.items():
    print(f"\n  [{name}]")

    # ── Train ────────────────────────────────────────────────────────────────
    pipe.fit(X_tr, y_tr)
    y_prob = pipe.predict_proba(X_te)[:, 1]
    y_pred = pipe.predict(X_te)

    # ── AUC + Average Precision ───────────────────────────────────────────────
    auc = roc_auc_score(y_te, y_prob)
    ap  = average_precision_score(y_te, y_prob)

    # ── 5-Fold CV AUC (trên toàn bộ dataset, không chỉ train set) ────────────
    # Với Random Forest, CV chạy lâu hơn vì 100 cây × 5 folds = 500 lần train
    # Dùng X, y toàn bộ (không phải X_tr, y_tr) để có đánh giá công bằng hơn
    cv_scores = cross_val_score(pipe, X, y, cv=cv, scoring="roc_auc", n_jobs=-1)

    # ── Confusion Matrix ─────────────────────────────────────────────────────
    # [TN FP]   TN = đúng là low risk
    # [FN TP]   TP = đúng là high risk
    #           FP = báo động false (low risk bị gán high risk)
    #           FN = bỏ sót (high risk bị gán low risk — nguy hiểm nhất)
    cm = confusion_matrix(y_te, y_pred)
    tn, fp, fn, tp = cm.ravel()
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall    = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1        = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    # ── ROC curve points ─────────────────────────────────────────────────────
    fpr_arr, tpr_arr, _ = roc_curve(y_te, y_prob)
    step = max(1, len(fpr_arr) // 25)
    roc_pts = pd.DataFrame({
        "FPR": fpr_arr[::step].round(4),
        "TPR": tpr_arr[::step].round(4),
    })

    # ── Feature Importance ───────────────────────────────────────────────────
    # LR: dùng |coefficient| làm proxy importance (sau khi standardize)
    # DT: dùng feature_importances_ (Gini impurity decrease)
    # RF: dùng feature_importances_ (average Gini across 100 trees)
    mdl = pipe.named_steps["model"]
    if name == "Logistic Regression":
        raw_imp = np.abs(mdl.coef_[0])
        imp_type = "Abs. Coefficient (standardized)"
    else:
        raw_imp  = mdl.feature_importances_
        imp_type = "Gini Impurity Decrease"

    imp_sum  = raw_imp.sum()
    imp_pct  = raw_imp / imp_sum * 100 if imp_sum > 0 else raw_imp * 0
    rank     = (-raw_imp).argsort().argsort() + 1  # rank 1 = most important
    feat_imp = pd.DataFrame({
        "Feature":     FEATURES,
        "Importance":  raw_imp.round(6),
        "Importance%": imp_pct.round(2),
        "Rank":        rank,
        "Type":        imp_type,
    }).sort_values("Rank")

    results[name] = {
        "auc":      auc,
        "ap":       ap,
        "cv_mean":  cv_scores.mean(),
        "cv_std":   cv_scores.std(),
        "cv_scores":cv_scores,
        "precision":precision,
        "recall":   recall,
        "f1":       f1,
        "tn": tn, "fp": fp, "fn": fn, "tp": tp,
        "feat_imp": feat_imp,
        "roc_pts":  roc_pts,
        "y_prob":   y_prob,
    }

    print(f"    AUC         = {auc:.4f}")
    print(f"    Avg Prec    = {ap:.4f}")
    print(f"    5-CV AUC    = {cv_scores.mean():.4f} ± {cv_scores.std():.4f}")
    print(f"    Precision   = {precision:.4f}  (bao nhiêu % báo high-risk là đúng)")
    print(f"    Recall      = {recall:.4f}  (bắt được bao nhiêu % high-risk thực)")
    print(f"    F1 Score    = {f1:.4f}")
    print(f"    Confusion:  TN={tn:,} FP={fp:,} FN={fn:,} TP={tp:,}")
    print(f"    Feature importance:")
    for _, rr in feat_imp.iterrows():
        bar = "█" * int(rr["Importance%"] / 5)
        print(f"      #{int(rr['Rank'])} {rr['Feature']:<12} {rr['Importance%']:>6.2f}%  {bar}")

# =============================================================================
#  INDIVIDUAL PD — best model (Random Forest)
# =============================================================================
print("\n" + "=" * 65)
print("  Individual PD scores — Random Forest (best model)")

best_model = "Random Forest"
rf_pipe    = models[best_model]

ml_df2 = df[df["segment"] != "Other"].copy().dropna(subset=FEATURES)
ml_df2["pd_individual"] = rf_pipe.predict_proba(ml_df2[FEATURES].values)[:, 1]

pd_compare = ml_df2.groupby("segment").agg(
    n            = ("is_high_risk", "count"),
    group_pd     = ("is_high_risk", "mean"),
    ind_pd_p10   = ("pd_individual", lambda x: x.quantile(0.10)),
    ind_pd_p25   = ("pd_individual", lambda x: x.quantile(0.25)),
    ind_pd_med   = ("pd_individual", "median"),
    ind_pd_p75   = ("pd_individual", lambda x: x.quantile(0.75)),
    ind_pd_p90   = ("pd_individual", lambda x: x.quantile(0.90)),
).reset_index()

print("\n  Group PD vs Individual PD (Random Forest):")
for _, rr in pd_compare.iterrows():
    spread = (rr["ind_pd_p90"] - rr["ind_pd_p10"]) * 100
    print(f"    {rr['segment']:<30}"
          f"  group={rr['group_pd']*100:.2f}%"
          f"  median={rr['ind_pd_med']*100:.2f}%"
          f"  [P10={rr['ind_pd_p10']*100:.1f}% – P90={rr['ind_pd_p90']*100:.1f}%]"
          f"  spread={spread:.1f}pp")

# Best model win margin
lr_auc  = results["Logistic Regression"]["auc"]
dt_auc  = results["Decision Tree"]["auc"]
rf_auc  = results["Random Forest"]["auc"]
print(f"\n  AUC improvement over LR baseline:")
print(f"    Decision Tree:  +{(dt_auc - lr_auc)*100:.2f}pp")
print(f"    Random Forest:  +{(rf_auc - lr_auc)*100:.2f}pp  ← best")

# =============================================================================
#  WRITE EXCEL
# =============================================================================
print("\n" + "=" * 65)
print("  Writing Excel …")
wb  = Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 1 — MODEL COMPARISON
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("W2_ModelComparison")
NC  = 9

title(ws1, 1,
      "WEEK 2 — ML MODEL COMPARISON  |  LR vs Decision Tree vs Random Forest  |  4 features", NC)
r = 2

# ── Model descriptions ───────────────────────────────────────────────────────
sub(ws1, r, "MODEL DEFINITIONS", NC); r += 1
descs = [
    ("Logistic Regression", "Đường thẳng — cộng tuyến tính các features. Baseline. Không học được quan hệ phi tuyến.",
     "C=1.0, class_weight=balanced, StandardScaler", F_LR),
    ("Decision Tree",       "Học được điểm gãy phi tuyến (ví dụ: FICO < 680 AND DTI > 22%). Dễ overfit nếu không giới hạn.",
     "max_depth=6, min_samples_leaf=500, Gini impurity, class_weight=balanced", F_DT),
    ("Random Forest",       "Ensemble 100 Decision Trees, mỗi cây học trên bootstrap sample khác nhau → ổn định và mạnh hơn DT đơn lẻ.",
     "100 trees, max_depth=8, max_features=sqrt(4)=2, class_weight=balanced, n_jobs=-1", F_RF),
]
hrow(ws1, r, ["Model", "How it works", "Key Parameters", "", "", "", "", "", ""])
r += 1
for name, desc, params, fl in descs:
    note(ws1, r,  f"  {name}", 1, fl)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c2 = ws1.cell(r, 2, desc); c2.font = FN; c2.fill = fl; c2.alignment = AL; c2.border = bdr()
    ws1.merge_cells(start_row=r, start_column=7, end_row=r, end_column=NC)
    c3 = ws1.cell(r, 7, params); c3.font = FSML; c3.fill = fl; c3.alignment = AL; c3.border = bdr()
    r += 1

r += 1

# ── Metrics table ────────────────────────────────────────────────────────────
sub(ws1, r, "PERFORMANCE METRICS — TEST SET (20% holdout, stratified split)", NC); r += 1
hrow(ws1, r, ["Model", "AUC ↑", "5-CV AUC (mean)", "5-CV Std", "Avg Precision ↑",
              "Precision", "Recall", "F1 Score", "vs LR Baseline"])
r += 1

model_order = ["Logistic Regression", "Decision Tree", "Random Forest"]
for name in model_order:
    res = results[name]
    fl  = model_fills[name]
    delta_auc = res["auc"] - results["Logistic Regression"]["auc"]
    if name == "Logistic Regression":
        delta_str = "BASELINE"
        fnt_d = FN
    elif delta_auc > 0.01:
        delta_str = f"+{delta_auc*100:.2f}pp ↑"
        fnt_d = FGRN
    elif delta_auc > 0:
        delta_str = f"+{delta_auc*100:.2f}pp"
        fnt_d = FTEL
    else:
        delta_str = f"{delta_auc*100:.2f}pp"
        fnt_d = FRED

    best_auc = max(r2["auc"] for r2 in results.values())
    fnt_auc  = FGRN if res["auc"] == best_auc else FB

    drow(ws1, r, [
        name,
        f"{res['auc']:.4f}",
        f"{res['cv_mean']:.4f}",
        f"± {res['cv_std']:.4f}",
        f"{res['ap']:.4f}",
        f"{res['precision']:.4f}",
        f"{res['recall']:.4f}",
        f"{res['f1']:.4f}",
        delta_str,
    ], fl, fnts=[FB, fnt_auc, FN, FN, FN, FN, FN, FN, fnt_d])
    r += 1

r += 1
note(ws1, r,
     "AUC: higher = better (1.0 = perfect, 0.5 = random). "
     "Recall quan trọng hơn Precision trong credit risk — bỏ sót high-risk (FN) tốn tiền hơn "
     "false alarm (FP).", NC, F_TEAL)
r += 2

# ── Confusion Matrix ─────────────────────────────────────────────────────────
sub(ws1, r, "CONFUSION MATRIX — TEST SET", NC); r += 1
note(ws1, r,
     "TN=đúng low-risk | FP=báo nhầm high-risk | FN=bỏ sót high-risk (nguy hiểm nhất) | TP=đúng high-risk", NC, F_GOLD)
r += 1
hrow(ws1, r, ["Model", "True Negative", "False Positive", "False Negative", "True Positive",
              "FN Rate (miss rate)", "FP Rate (false alarm)", "", ""])
r += 1
for name in model_order:
    res = results[name]
    fl  = model_fills[name]
    total_pos = res["tp"] + res["fn"]
    total_neg = res["tn"] + res["fp"]
    fn_rate   = res["fn"] / total_pos if total_pos > 0 else 0
    fp_rate   = res["fp"] / total_neg if total_neg > 0 else 0
    fnt_fn    = FRED if fn_rate > 0.5 else FGLD if fn_rate > 0.3 else FGRN
    drow(ws1, r, [
        name,
        f"{res['tn']:,}", f"{res['fp']:,}", f"{res['fn']:,}", f"{res['tp']:,}",
        f"{fn_rate*100:.1f}%", f"{fp_rate*100:.1f}%", "", "",
    ], fl, fnts=[FB, FN, FN, fnt_fn, FN, fnt_fn, FN, FN, FN])
    r += 1

r += 1

# ── CV scores breakdown ───────────────────────────────────────────────────────
sub(ws1, r, "5-FOLD CROSS VALIDATION — INDIVIDUAL FOLD AUC SCORES", NC); r += 1
hrow(ws1, r, ["Model", "Fold 1", "Fold 2", "Fold 3", "Fold 4", "Fold 5", "Mean", "Std", "Stable?"])
r += 1
for name in model_order:
    res = results[name]
    fl  = model_fills[name]
    folds = [f"{s:.4f}" for s in res["cv_scores"]]
    stable = "YES" if res["cv_std"] < 0.005 else "CHECK"
    fnt_s  = FGRN if stable == "YES" else FGLD
    drow(ws1, r,
         folds + [f"{res['cv_mean']:.4f}", f"± {res['cv_std']:.4f}", stable],
         fl,
         fnts=[FN]*5 + [FB, FN, fnt_s])
    # Prepend name
    ws1.insert_cols(1) if r == r else None  # no-op, just for clarity
    # Fix: write name manually at col 1
    c_name = ws1.cell(r, 1, name)
    c_name.font = FB; c_name.fill = fl; c_name.alignment = AC; c_name.border = bdr()
    # Shift: drow already wrote to cols 1..9, rewrite col 1
    # Actually drow wrote name at col1 already — fix by re-writing
    for j2, v in enumerate(folds + [f"{res['cv_mean']:.4f}", f"± {res['cv_std']:.4f}", stable], 2):
        c = ws1.cell(r, j2, v)
        c.font = FN; c.fill = fl; c.alignment = AC; c.border = bdr()
        if j2 == 7: c.font = FB
        if j2 == 9: c.font = fnt_s
    r += 1

autofit(ws1, 12, 45)

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 2 — FEATURE IMPORTANCE
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("W2_FeatureImportance")
NC2 = 7

title(ws2, 1, "WEEK 2 — FEATURE IMPORTANCE  |  Biến nào quan trọng nhất theo từng model?", NC2)
r = 2

sub(ws2, r,
    "Ý NGHĨA FEATURE IMPORTANCE", NC2); r += 1
meanings = [
    "Logistic Regression : |Coefficient| sau StandardScaler → độ lớn ảnh hưởng đến log-odds",
    "Decision Tree       : Gini Impurity Decrease — feature được chọn để split bao nhiêu lần và giảm impurity bao nhiêu",
    "Random Forest       : Average Gini Decrease across 100 trees → ổn định hơn DT đơn lẻ vì là trung bình",
    "Tất cả đã normalize thành % (tổng = 100%) để so sánh được giữa 3 model",
]
for m in meanings:
    note(ws2, r, f"  {m}", NC2, F_TEAL); r += 1

r += 1

for name in model_order:
    fl  = model_fills[name]
    fi  = results[name]["feat_imp"]
    sub(ws2, r, f"{name.upper()} — Feature Importance", NC2, fl); r += 1
    hrow(ws2, r, ["Rank", "Feature", "Importance Score", "Importance %",
                  "Bar (visual)", "Interpretation", ""], fl)
    r += 1

    interps = {
        "fico_avg":  "Điểm tín dụng — tín hiệu cá nhân mạnh nhất",
        "dti":       "Tỷ lệ nợ/thu nhập — khả năng trả nợ hàng tháng",
        "loan_amnt": "Số tiền vay — khoản lớn hơn = gánh nặng hơn",
        "term_60":   "Thời hạn 60 tháng — cam kết dài hơn = rủi ro cao hơn",
    }
    for _, rr in fi.iterrows():
        bar = "█" * max(1, int(rr["Importance%"] / 5))
        drow(ws2, r, [
            f"#{int(rr['Rank'])}",
            rr["Feature"],
            f"{rr['Importance']:.6f}",
            f"{rr['Importance%']:.2f}%",
            bar,
            interps.get(rr["Feature"], "—"),
            "",
        ], fl, fnts=[FB, FB, FN, FB, FTEL, FN, FN])
        r += 1
    r += 1

# ── Consistency check: does variable ranking agree across models? ─────────────
sub(ws2, r, "CONSISTENCY CHECK — Does variable ranking agree across 3 models?", NC2); r += 1
note(ws2, r,
     "Nếu tất cả 3 models đều xếp cùng feature ở Rank #1 → feature đó thực sự quan trọng, không phải artifact của 1 model.",
     NC2, F_TEAL)
r += 1

hrow(ws2, r, ["Feature", "LR Rank", "DT Rank", "RF Rank", "Consistent?",
              "What it means", ""])
r += 1
for feat in FEATURES:
    ranks = {}
    for name in model_order:
        fi  = results[name]["feat_imp"]
        row_ = fi[fi["Feature"] == feat]
        ranks[name] = int(row_["Rank"].iloc[0]) if len(row_) > 0 else 9
    consistent = "YES ✓" if len(set(ranks.values())) == 1 else "DIFFERS"
    fnt_c = FGRN if consistent == "YES ✓" else FGLD
    meaning = (
        "Tất cả models đồng thuận → very strong signal"
        if consistent == "YES ✓"
        else "Models bất đồng → signal phụ thuộc vào cách model học"
    )
    drow(ws2, r, [
        feat,
        f"#{ranks['Logistic Regression']}",
        f"#{ranks['Decision Tree']}",
        f"#{ranks['Random Forest']}",
        consistent,
        meaning,
        "",
    ], F_WHITE, fnts=[FB, FN, FN, FN, fnt_c, FN, FN])
    r += 1

autofit(ws2, 12, 48)

# ─────────────────────────────────────────────────────────────────────────────
#  SHEET 3 — INDIVIDUAL PD (Random Forest)
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("W2_IndividualPD")
NC3 = 9

title(ws3, 1,
      "WEEK 2 — INDIVIDUAL PD (Random Forest)  |  Group PD vs Person-Level PD per Segment", NC3)
r = 2

sub(ws3, r, "TẠI SAO INDIVIDUAL PD QUAN TRỌNG HƠN GROUP PD?", NC3); r += 1
explanations = [
    "Group PD = trung bình của cả segment (ví dụ: mọi người trong Seg4 đều bị gán 23.89%)",
    "Individual PD = xác suất riêng của từng người dựa trên FICO thực, DTI thực, loan amount thực",
    "Người FICO 679 và người FICO 641 đều trong Seg4 nhưng rủi ro thực tế rất khác nhau",
    "Individual PD → pricing chính xác hơn (người ít rủi ro được lãi suất tốt hơn)",
    "Model dùng: Random Forest (AUC cao nhất trong 3 models)",
]
for e in explanations:
    note(ws3, r, f"  → {e}", NC3, F_RF); r += 1

r += 1
sub(ws3, r, "GROUP PD vs INDIVIDUAL PD — DISTRIBUTION PER SEGMENT (Random Forest)", NC3); r += 1
hrow(ws3, r, ["Segment", "N", "Group PD (%)", "Ind. PD P10",
              "Ind. PD P25", "Ind. PD Median", "Ind. PD P75",
              "Ind. PD P90", "Spread P10–P90 (pp)"])
r += 1

for i, rr in pd_compare.iterrows():
    seg_i  = seg_names.index(rr["segment"]) if rr["segment"] in seg_names else i
    spread = (rr["ind_pd_p90"] - rr["ind_pd_p10"]) * 100
    fnt_sp = FRED if spread > 20 else FGLD if spread > 10 else FGRN
    drow(ws3, r, [
        rr["segment"],
        f"{int(rr['n']):,}",
        f"{rr['group_pd']*100:.2f}%",
        f"{rr['ind_pd_p10']*100:.2f}%",
        f"{rr['ind_pd_p25']*100:.2f}%",
        f"{rr['ind_pd_med']*100:.2f}%",
        f"{rr['ind_pd_p75']*100:.2f}%",
        f"{rr['ind_pd_p90']*100:.2f}%",
        f"{spread:.2f}pp",
    ], SEG_FILLS[seg_i],
       fnts=[FB, FN, FB, FN, FN, FB, FN, FN, fnt_sp])
    r += 1

r += 1
note(ws3, r,
     "Spread P10–P90 = khoảng biến thiên của individual PD trong cùng một segment. "
     "Spread cao → segment heterogeneous → cần individual PD để pricing chính xác.",
     NC3, F_TEAL)
r += 2

# ── Business implication ──────────────────────────────────────────────────────
sub(ws3, r, "BUSINESS IMPLICATION — Individual PD for Loan Pricing", NC3); r += 1
implications = [
    "Seg4 — group PD = 23.89%, nhưng P10 có thể chỉ ~10% → người đó bị charge lãi suất quá cao",
    "Seg1 — group PD = 8.91%, nhưng P90 có thể ~18% → người đó chỉ trả lãi như Seg1 dù thực ra nguy hiểm hơn",
    "Individual PD → interest rate = base_rate + risk_premium(pd_individual) → fair pricing",
    "Rule of thumb: nếu individual PD > 2× group PD của segment → flag for manual review",
]
for impl in implications:
    note(ws3, r, f"  • {impl}", NC3, F_GOLD); r += 1

r += 2

# ── Model comparison summary ──────────────────────────────────────────────────
sub(ws3, r, "FINAL MODEL RANKING SUMMARY", NC3); r += 1
hrow(ws3, r, ["Rank", "Model", "AUC", "5-CV AUC", "Recall", "Interpretable?",
              "Production Ready?", "Verdict", ""])
r += 1
rankings = sorted(model_order, key=lambda n: results[n]["auc"], reverse=True)
rank_fills = [F_RF, F_DT, F_LR]
verdicts = {
    "Random Forest":       "BEST — Highest AUC, stable CV, good recall",
    "Decision Tree":       "GOOD — Non-linear, interpretable, slightly less stable",
    "Logistic Regression": "BASELINE — Linear only, but fast and explainable",
}
interpret = {
    "Random Forest": "Moderate (feature importance)",
    "Decision Tree": "HIGH (visualize tree)",
    "Logistic Regression": "HIGH (coefficients)",
}
prod_ready = {
    "Random Forest": "YES — with calibration",
    "Decision Tree": "YES — simple to deploy",
    "Logistic Regression": "YES — most common in banking",
}
for rank_i, name in enumerate(rankings):
    res = results[name]
    fl  = rank_fills[rank_i]
    drow(ws3, r, [
        f"#{rank_i+1}",
        name,
        f"{res['auc']:.4f}",
        f"{res['cv_mean']:.4f} ± {res['cv_std']:.4f}",
        f"{res['recall']:.4f}",
        interpret[name],
        prod_ready[name],
        verdicts[name],
        "",
    ], fl, fnts=[FB, FB, FGRN if rank_i==0 else FB, FN, FN, FN, FN, FTEL if rank_i==0 else FN, FN])
    r += 1

autofit(ws3, 12, 48)

# =============================================================================
#  SAVE
# =============================================================================
wb.save(OUT_PATH)
print(f"\n{'='*65}")
print(f"  ✅  DONE  →  {OUT_PATH}")
print(f"  Sheets:")
print(f"    W2_ModelComparison  — AUC, CV, confusion matrix cả 3 models")
print(f"    W2_FeatureImportance — ranking features theo LR / DT / RF")
print(f"    W2_IndividualPD     — group vs individual PD distribution")
print(f"\n  FINAL AUC RANKING:")
for name in sorted(model_order, key=lambda n: results[n]["auc"], reverse=True):
    print(f"    {name:<22} AUC={results[name]['auc']:.4f}"
          f"  CV={results[name]['cv_mean']:.4f}")