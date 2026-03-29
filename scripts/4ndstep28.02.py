# final_analysis.py
# Làm 4 việc trong 1 lần chạy:
#   Sheet 1 — Total Exposure (câu 2)
#   Sheet 2 — Biggest Risk segment (câu 1)
#   Sheet 3 — Business Decisions (câu 3)
#   Sheet 4 — Variable Importance (câu 4)

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
CSV_PATH = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\week6_clean_data.csv"
OUT_PATH = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\final_analysis.xlsx"

FICO_CUT  = 680.0
LGD       = 0.65   # Loss Given Default — flat fallback (used if 'recoveries' column missing)

# =========================
# CÔNG THỨC TÍNH LGD THỰC TẾ
# =========================
# LGD (actual) = 1 - Recovery Rate
# Recovery Rate = recoveries / funded_amnt
#
# Chỉ tính cho high_risk loans (Charged Off, Default, Late 31-120d)
# Nếu recoveries = $2,000 và funded_amnt = $10,000
#   → Recovery Rate = 20%  →  LGD = 80%
#
# Hàm này trả về dict {segment: lgd_value}
# Được gọi SAU khi df đã load và segment đã được gán nhãn
def compute_actual_lgd(df_):
    """
    Tính LGD thực tế từ cột 'recoveries' và 'funded_amnt'.
    Trả về dict segment → lgd_float (fallback = 0.65 nếu thiếu data).
    """
    if "recoveries" not in df_.columns:
        print("  ⚠️  'recoveries' not found — using flat LGD = 65%")
        return {}

    denom = "funded_amnt" if "funded_amnt" in df_.columns else "loan_amnt"
    hr    = df_[df_["risk_label"] == "high_risk"].copy()

    # Recovery Rate per loan, capped [0, 1]
    hr["_rec_rate"] = (
        hr["recoveries"]
        .div(hr[denom].replace(0, float("nan")))
        .clip(0, 1)
        .fillna(0)
    )

    result = {}
    for seg, grp in hr.groupby("segment"):
        if seg == "Other":
            continue
        avg_rec = grp["_rec_rate"].mean()
        result[seg] = round(1.0 - avg_rec, 4)
        print(f"    LGD actual [{seg}] = {result[seg]*100:.2f}%  "
              f"(recovery rate = {avg_rec*100:.2f}%,  n_hr = {len(grp):,})")

    return result

# =========================
# STYLES
# =========================
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

FILL_HEADER  = fill("1E3A5F")
FILL_SEG1    = fill("C6EFCE")
FILL_SEG2    = fill("FFEB9C")
FILL_SEG3    = fill("FCE4D6")
FILL_SEG4    = fill("FF9999")
FILL_TITLE   = fill("0F2942")
FILL_BEST    = fill("C6EFCE")
FILL_WORST   = fill("FF9999")
FILL_MID     = fill("FFEB9C")
FILL_SUBHEAD = fill("2D5F8A")

SEG_FILLS = [FILL_SEG1, FILL_SEG2, FILL_SEG3, FILL_SEG4]

FONT_TITLE  = Font(bold=True, size=14, color="FFFFFF")
FONT_HEADER = Font(bold=True, size=10, color="FFFFFF")
FONT_SUBHD  = Font(bold=True, size=10, color="FFFFFF")
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
ALIGN_C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, ncols, font=FONT_HEADER, fill_=FILL_HEADER):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font   = font
        cell.fill   = fill_
        cell.alignment = ALIGN_C
        cell.border = thin_border()

def style_data_row(ws, row, ncols, fill_=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font      = FONT_NORMAL
        cell.alignment = ALIGN_C
        cell.border    = thin_border()
        if fill_:
            cell.fill = fill_

def autofit(ws, min_w=12, max_w=40):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        col_letter = next((c.column_letter for c in col if not isinstance(c, MergedCell)), None)
        if col_letter is None:
            continue
        max_len = max((len(str(c.value)) for c in col if c.value and not isinstance(c, MergedCell)), default=10)
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)

def write_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = FONT_TITLE; c.fill = FILL_TITLE; c.alignment = ALIGN_C; c.border = thin_border()

def write_subheader(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = FONT_SUBHD; c.fill = FILL_SUBHEAD; c.alignment = ALIGN_L; c.border = thin_border()

# =========================
# LOAD DATA
# =========================
print("Loading data...")
df = pd.read_csv(CSV_PATH, low_memory=False)
df["fico_binary"] = (df["fico_avg"] >= FICO_CUT).map({True: "high_fico", False: "low_fico"})

baseline = (df["risk_label"] == "high_risk").mean() * 100
total_n  = len(df)
print(f"  Rows: {total_n:,} | Baseline: {baseline:.2f}%")

# =========================
# HELPER: compute risk table for any groupby
# =========================
def risk_table(df_, group_cols):
    grp = df_.groupby(group_cols)
    out = grp.agg(
        n          = ("risk_label", "count"),
        high_risk  = ("risk_label", lambda x: (x == "high_risk").sum()),
        avg_loan   = ("loan_amnt",  "mean"),
    ).reset_index()
    out["high_risk_pct"] = out["high_risk"] / out["n"] * 100
    out["delta_pp"]      = out["high_risk_pct"] - baseline
    out["lift"]          = out["high_risk_pct"] / baseline
    return out

# =========================
# SEGMENT DEFINITIONS
# =========================
def assign_segment(df_):
    cond = [
        (df_["fico_avg"] >= FICO_CUT) & (df_["term_m"] == 36),
        (df_["fico_avg"] >= FICO_CUT) & (df_["term_m"] == 60),
        (df_["fico_avg"] <  FICO_CUT) & (df_["term_m"] == 36),
        (df_["fico_avg"] <  FICO_CUT) & (df_["term_m"] == 60),
    ]
    names = ["Seg1 — Prime Short-Term",
             "Seg2 — Prime Long-Term",
             "Seg3 — Subprime Short-Term",
             "Seg4 — Subprime Long-Term"]
    return np.select(cond, names, default="Other")

df["segment"] = assign_segment(df)

# =========================
# TÍNH LGD THỰC TẾ TỪ DATA
# =========================
print("\nComputing actual LGD from data...")
actual_lgd_map = compute_actual_lgd(df)
# Nếu không tính được (thiếu cột) → dùng flat 65% cho tất cả

seg_stats = risk_table(df[df["segment"] != "Other"], ["segment"])
seg_stats = seg_stats.sort_values("segment").reset_index(drop=True)

# Add medium risk
med = df[df["segment"] != "Other"].groupby("segment").apply(
    lambda x: (x["risk_label"] == "medium_risk").mean() * 100
).reset_index(name="med_risk_pct")
seg_stats = seg_stats.merge(med, on="segment")

# Gán LGD cho từng segment (actual nếu có, fallback = 0.65)
seg_stats["lgd_used"] = seg_stats["segment"].map(
    lambda s: actual_lgd_map.get(s, LGD)
)

# Total Exposure (EL = PD × LGD × EAD)
# Trước: dùng LGD = 0.65 flat
# Sau:   dùng lgd_used — actual LGD từ cột recoveries nếu có
seg_stats["total_exposure"] = (
    seg_stats["n"] *
    seg_stats["avg_loan"] *
    (seg_stats["high_risk_pct"] / 100) *
    seg_stats["lgd_used"]          # ← THAY ĐỔI: dùng actual LGD thay flat LGD
)
seg_stats["pct_of_total"] = seg_stats["n"] / total_n * 100

SEG_NAMES = {
    "Seg1 — Prime Short-Term":    "Prime Short-Term",
    "Seg2 — Prime Long-Term":     "Prime Long-Term",
    "Seg3 — Subprime Short-Term": "Subprime Short-Term",
    "Seg4 — Subprime Long-Term":  "Subprime Long-Term",
}

print("\nSegment stats:")
for _, r in seg_stats.iterrows():
    print(f"  {r['segment']}: n={r['n']:,} | HR={r['high_risk_pct']:.2f}% | Exposure=${r['total_exposure']/1e6:.1f}M")

# =========================
# VARIABLE IMPORTANCE
# =========================
print("\nVariable importance analysis...")

def gap(df_, col1, col2):
    t = risk_table(df_, [col1, col2])
    return t["high_risk_pct"].max() - t["high_risk_pct"].min()

def gap_3way(df_):
    t = risk_table(df_, ["fico_binary", "dti_group", "term_group"])
    return t["high_risk_pct"].max() - t["high_risk_pct"].min()

gap_full      = gap_3way(df)
gap_no_dti    = gap(df, "fico_binary", "term_group")
gap_no_term   = gap(df, "fico_binary", "dti_group")
gap_no_fico   = gap(df, "dti_group",   "term_group")

varimp = pd.DataFrame([
    {"Scenario": "Full (FICO + DTI + Term)", "Variables Used": "All 3",
     "Worst %": None, "Best %": None, "Gap (pp)": gap_full, "Gap Lost (pp)": 0,
     "% Gap Lost": 0, "Most Critical Removed": "—"},
    {"Scenario": "Drop DTI → FICO + Term", "Variables Used": "FICO, Term",
     "Gap (pp)": gap_no_dti,
     "Gap Lost (pp)": gap_full - gap_no_dti,
     "% Gap Lost": (gap_full - gap_no_dti) / gap_full * 100,
     "Most Critical Removed": "DTI"},
    {"Scenario": "Drop Term → FICO + DTI", "Variables Used": "FICO, DTI",
     "Gap (pp)": gap_no_term,
     "Gap Lost (pp)": gap_full - gap_no_term,
     "% Gap Lost": (gap_full - gap_no_term) / gap_full * 100,
     "Most Critical Removed": "Term"},
    {"Scenario": "Drop FICO → DTI + Term", "Variables Used": "DTI, Term",
     "Gap (pp)": gap_no_fico,
     "Gap Lost (pp)": gap_full - gap_no_fico,
     "% Gap Lost": (gap_full - gap_no_fico) / gap_full * 100,
     "Most Critical Removed": "FICO"},
])

# Find worst gap lost = most critical variable
most_critical = varimp.iloc[1:]["Gap Lost (pp)"].idxmax()
print(f"\nVariable Importance gaps:")
for _, r in varimp.iterrows():
    print(f"  {r['Scenario']}: gap={r['Gap (pp)']:.2f}pp | lost={r['Gap Lost (pp)']:.2f}pp")

# =========================
# BUSINESS DECISIONS
# =========================
biz = [
    {
        "Segment": "Seg1 — Prime Short-Term",
        "Risk Profile": f"High Risk: {seg_stats.iloc[0]['high_risk_pct']:.2f}% | Lift: {seg_stats.iloc[0]['lift']:.2f}x",
        "Action": "Fast-track approval + interest rate discount",
        "Rationale": "Lowest risk in portfolio. FICO good + short term = most reliable borrowers. Priority: grow this segment.",
        "Monitoring": "Quarterly review — maintain standards, no relaxation",
        "Pricing": "Best rate tier — reward low risk profile",
    },
    {
        "Segment": "Seg2 — Prime Long-Term",
        "Risk Profile": f"High Risk: {seg_stats.iloc[1]['high_risk_pct']:.2f}% | Lift: {seg_stats.iloc[1]['lift']:.2f}x",
        "Action": "Approve with income verification + encourage refinance to 36M",
        "Rationale": "FICO good but 5 years = hidden risk. Loan amount ~60% higher than Seg1 → chose 60M because loan is too large for 36M.",
        "Monitoring": "Monitor income & DTI semi-annually. Flag if DTI increases >5pp",
        "Pricing": "Standard rate + small premium for term risk",
    },
    {
        "Segment": "Seg3 — Subprime Short-Term",
        "Risk Profile": f"High Risk: {seg_stats.iloc[2]['high_risk_pct']:.2f}% | Lift: {seg_stats.iloc[2]['lift']:.2f}x",
        "Action": "Approve with loan amount cap + enforce 36M as mandatory",
        "Rationale": "FICO weak but 36M self-limits exposure. Risk same as Seg2 (0.13pp diff) → 36M is the protection mechanism. Do NOT allow 60M.",
        "Monitoring": "Monthly payment behavior tracking. Early warning if 1 missed payment",
        "Pricing": "Risk-adjusted rate — higher than Seg1/Seg2 to compensate PD",
    },
    {
        "Segment": "Seg4 — Subprime Long-Term",
        "Risk Profile": f"High Risk: {seg_stats.iloc[3]['high_risk_pct']:.2f}% | Lift: {seg_stats.iloc[3]['lift']:.2f}x",
        "Action": "Restrict or decline. If approve: high risk premium + reduced loan cap",
        "Rationale": "Both signals bad simultaneously. ~1 in 4 will default. Highest exposure risk. No compensating factor.",
        "Monitoring": "Quarterly early warning. Flag at first missed payment for workout process",
        "Pricing": "Maximum risk premium rate or decline based on specific FICO + DTI combination",
    },
]

# =========================
# WRITE EXCEL
# =========================
print("\nWriting Excel...")
wb = Workbook()

# -----------------------------------------------
# SHEET 1 — TOTAL EXPOSURE
# -----------------------------------------------
ws1 = wb.active
ws1.title = "Total_Exposure"
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20

write_title(ws1, 1, f"TOTAL EXPOSURE ANALYSIS  |  LGD = {LGD:.0%}  |  Baseline = {baseline:.2f}%  |  Formula: N × Avg Loan × High Risk % × LGD", 11)

headers = ["Segment", "Name", "N (loans)", "% of Total", "Avg FICO",
           "Avg Loan Amt", "High Risk %", "Δ vs Baseline", "Lift",
           "Total Exposure ($M)", "% of Total Exposure"]
for j, h in enumerate(headers, 1):
    ws1.cell(2, j, h)
style_header_row(ws1, 2, len(headers))

total_exp = seg_stats["total_exposure"].sum()
avg_ficos = df[df["segment"] != "Other"].groupby("segment")["fico_avg"].mean()

for i, row in seg_stats.iterrows():
    r = i + 3
    avg_f = avg_ficos.get(row["segment"], 0)
    vals = [
        row["segment"],
        SEG_NAMES.get(row["segment"], ""),
        f"{row['n']:,}",
        f"{row['pct_of_total']:.2f}%",
        f"{avg_f:.1f}",
        f"${row['avg_loan']:,.0f}",
        f"{row['high_risk_pct']:.2f}%",
        f"{row['delta_pp']:+.2f}pp",
        f"{row['lift']:.2f}x",
        f"${row['total_exposure']/1e6:.1f}M",
        f"{row['total_exposure']/total_exp*100:.1f}%",
    ]
    for j, v in enumerate(vals, 1):
        ws1.cell(r, j, v)
    style_data_row(ws1, r, len(headers), SEG_FILLS[i])

# Total row
r = len(seg_stats) + 3
ws1.cell(r, 1, "TOTAL")
ws1.cell(r, 3, f"{seg_stats['n'].sum():,}")
ws1.cell(r, 10, f"${total_exp/1e6:.1f}M")
ws1.cell(r, 11, "100%")
style_data_row(ws1, r, len(headers), fill("D9D9D9"))
for c in range(1, len(headers)+1):
    ws1.cell(r, c).font = FONT_BOLD

# Insight box
r += 2
write_subheader(ws1, r, "KEY INSIGHTS", len(headers))
r += 1
insights = [
    f"Highest Total Exposure: {seg_stats.loc[seg_stats['total_exposure'].idxmax(), 'segment']}  (${seg_stats['total_exposure'].max()/1e6:.1f}M — {seg_stats['total_exposure'].max()/total_exp*100:.1f}% of total)",
    f"Highest High Risk %:    {seg_stats.loc[seg_stats['high_risk_pct'].idxmax(), 'segment']}  ({seg_stats['high_risk_pct'].max():.2f}%)",
    f"Are they the same segment? {'YES ✅' if seg_stats['total_exposure'].idxmax() == seg_stats['high_risk_pct'].idxmax() else 'NO ⚠️ — highest % ≠ highest $ exposure'}",
    f"Total portfolio exposure:   ${total_exp/1e6:.0f}M  (assuming LGD = {LGD:.0%})",
]
for ins in insights:
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    c = ws1.cell(r, 1, ins)
    c.font = Font(size=10, bold=False)
    c.alignment = ALIGN_L
    c.border = thin_border()
    r += 1

autofit(ws1)

# -----------------------------------------------
# SHEET 2 — BIGGEST FINANCIAL RISK (câu 1)
# -----------------------------------------------
ws2 = wb.create_sheet("Biggest_Risk_Analysis")
write_title(ws2, 1, "Q1 — WHICH SEGMENT CREATES THE BIGGEST FINANCIAL RISK TO LENDINGCLUB?", 6)

r = 2
write_subheader(ws2, r, "COMPARISON: High Risk % vs Total Exposure", 6)
r += 1

h2 = ["Segment", "High Risk %", "Rank by %", "Total Exposure ($M)", "Rank by $", "Conclusion"]
for j, h in enumerate(h2, 1):
    ws2.cell(r, j, h)
style_header_row(ws2, r, len(h2))
r += 1

rank_pct = seg_stats["high_risk_pct"].rank(ascending=False).astype(int)
rank_exp = seg_stats["total_exposure"].rank(ascending=False).astype(int)

for i, row in seg_stats.iterrows():
    concl = ""
    if rank_pct[i] == 1 and rank_exp[i] == 1:
        concl = "Highest BOTH % and $ ← most dangerous"
    elif rank_exp[i] == 1:
        concl = "Highest $ despite lower % ← scale effect"
    elif rank_pct[i] == 1:
        concl = "Highest % but lower $ ← small segment"
    else:
        concl = f"Rank #{rank_pct[i]} by %, Rank #{rank_exp[i]} by $"

    vals = [
        SEG_NAMES.get(row["segment"], row["segment"]),
        f"{row['high_risk_pct']:.2f}%",
        f"#{rank_pct[i]}",
        f"${row['total_exposure']/1e6:.1f}M",
        f"#{rank_exp[i]}",
        concl,
    ]
    for j, v in enumerate(vals, 1):
        ws2.cell(r, j, v)
    style_data_row(ws2, r, len(h2), SEG_FILLS[i])
    r += 1

r += 1
write_subheader(ws2, r, "ANALYSIS — WHY THIS SEGMENT IS THE BIGGEST RISK", 6)
r += 1

# Find biggest financial risk
biggest_exp_idx  = seg_stats["total_exposure"].idxmax()
biggest_pct_idx  = seg_stats["high_risk_pct"].idxmax()
biggest_exp_seg  = seg_stats.loc[biggest_exp_idx]
biggest_pct_seg  = seg_stats.loc[biggest_pct_idx]
same             = biggest_exp_idx == biggest_pct_idx

analysis_lines = [
    f"Highest High Risk %: {SEG_NAMES.get(biggest_pct_seg['segment'])} at {biggest_pct_seg['high_risk_pct']:.2f}% — nearly 1 in 4 borrowers will default.",
    f"Highest Total Exposure: {SEG_NAMES.get(biggest_exp_seg['segment'])} at ${biggest_exp_seg['total_exposure']/1e6:.1f}M.",
    "",
    f"{'These are the SAME segment — ' + SEG_NAMES.get(biggest_exp_seg['segment']) + ' dominates on both dimensions.' if same else 'These are DIFFERENT segments — the largest financial risk is not the highest default rate.'}",
    "",
    "Why exposure can differ from default rate:",
    f"  → A segment with lower % but 3x more loans can create more $ damage",
    f"  → Avg loan amount also matters: Seg2 avg ${seg_stats.iloc[1]['avg_loan']:,.0f} vs Seg3 avg ${seg_stats.iloc[2]['avg_loan']:,.0f}",
    f"  → Seg2 and Seg3 have nearly identical High Risk % ({seg_stats.iloc[1]['high_risk_pct']:.2f}% vs {seg_stats.iloc[2]['high_risk_pct']:.2f}%) but different N and loan size",
]

for line in analysis_lines:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws2.cell(r, 1, line)
    c.font = Font(size=10)
    c.alignment = ALIGN_L
    c.border = thin_border()
    r += 1

autofit(ws2)

# -----------------------------------------------
# SHEET 3 — BUSINESS DECISIONS (câu 3)
# -----------------------------------------------
ws3 = wb.create_sheet("Business_Decisions")
write_title(ws3, 1, "Q3 — BUSINESS DECISIONS PER SEGMENT  |  If you are LendingClub", 6)

r = 2
h3 = ["Segment", "Risk Profile", "Recommended Action", "Rationale", "Monitoring", "Pricing Strategy"]
for j, h in enumerate(h3, 1):
    ws3.cell(r, j, h)
style_header_row(ws3, r, len(h3))
r += 1

for i, b in enumerate(biz):
    vals = [b["Segment"], b["Risk Profile"], b["Action"], b["Rationale"], b["Monitoring"], b["Pricing"]]
    for j, v in enumerate(vals, 1):
        ws3.cell(r, j, v)
    style_data_row(ws3, r, len(h3), SEG_FILLS[i])
    ws3.row_dimensions[r].height = 60
    r += 1

r += 1
write_subheader(ws3, r, "OVERARCHING STRATEGY", 6)
r += 1
strategy = [
    "Core principle: Segment 1 (48% of portfolio) is the profit engine — protect and grow it.",
    "Risk concentration: Segment 4 (8.4%) carries disproportionate risk — restrict aggressively.",
    f"Key tension: Seg2 and Seg3 have near-identical risk ({seg_stats.iloc[1]['high_risk_pct']:.2f}% vs {seg_stats.iloc[2]['high_risk_pct']:.2f}%) but different levers — for Seg2, monitor income; for Seg3, enforce 36M term.",
    "Do NOT use DTI alone to approve/deny — it adds marginal separation compared to FICO + Term.",
]
for line in strategy:
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws3.cell(r, 1, line)
    c.font = Font(size=10)
    c.alignment = ALIGN_L
    c.border = thin_border()
    r += 1

autofit(ws3)
for col in ["C", "D", "E", "F"]:
    ws3.column_dimensions[col].width = 35

# -----------------------------------------------
# SHEET 4 — VARIABLE IMPORTANCE (câu 4)
# -----------------------------------------------
ws4 = wb.create_sheet("Variable_Importance")
write_title(ws4, 1, "Q4 — VARIABLE IMPORTANCE: Drop 1 variable at a time, measure risk separation loss", 7)

r = 2
write_subheader(ws4, r, "RISK SEPARATION GAP: Worst combo % − Best combo %", 7)
r += 1

h4 = ["Scenario", "Variables Used", "Gap (pp)", "Gap Lost vs Full (pp)", "% of Gap Lost", "Interpretation", "Verdict"]
for j, h in enumerate(h4, 1):
    ws4.cell(r, j, h)
style_header_row(ws4, r, len(h4))
r += 1

interps = [
    "Baseline — full separation with all 3 signals",
    f"Removing DTI loses {gap_full - gap_no_dti:.2f}pp of separation → DTI contribution",
    f"Removing Term loses {gap_full - gap_no_term:.2f}pp of separation → Term contribution",
    f"Removing FICO loses {gap_full - gap_no_fico:.2f}pp of separation → FICO contribution",
]

verdicts = ["FULL MODEL", "", "", ""]
gaps_lost = [0, gap_full - gap_no_dti, gap_full - gap_no_term, gap_full - gap_no_fico]
max_lost = max(gaps_lost[1:])

row_fills = [fill("D9E1F2"), fill("C6EFCE"), fill("FFEB9C"), fill("FF9999")]

for i, (_, vrow) in enumerate(varimp.iterrows()):
    lost = gaps_lost[i]
    pct_lost = lost / gap_full * 100 if gap_full > 0 else 0

    if i == 0:
        verdict = "BASELINE"
    elif lost == max_lost:
        verdict = "⚠️ MOST CRITICAL"
    elif lost == min(gaps_lost[1:]):
        verdict = "✅ LEAST CRITICAL"
    else:
        verdict = "MODERATE IMPACT"

    vals = [
        vrow["Scenario"],
        vrow["Variables Used"],
        f"{vrow['Gap (pp)']:.2f}pp",
        f"{lost:.2f}pp" if i > 0 else "—",
        f"{pct_lost:.1f}%" if i > 0 else "—",
        interps[i],
        verdict,
    ]
    for j, v in enumerate(vals, 1):
        ws4.cell(r, j, v)
    style_data_row(ws4, r, len(h4), row_fills[i])
    r += 1

r += 1
write_subheader(ws4, r, "RANKING — MOST TO LEAST CRITICAL VARIABLE", 7)
r += 1

# Sort by gap lost
var_ranking = [
    ("FICO",  gap_full - gap_no_fico,  (gap_full - gap_no_fico) / gap_full * 100),
    ("Term",  gap_full - gap_no_term,  (gap_full - gap_no_term) / gap_full * 100),
    ("DTI",   gap_full - gap_no_dti,   (gap_full - gap_no_dti)  / gap_full * 100),
]
var_ranking.sort(key=lambda x: -x[1])

rank_fills = [FILL_WORST, FILL_MID, FILL_BEST]
rank_heads = ["Rank", "Variable", "Gap Lost if Removed (pp)", "% of Total Gap Lost", "Conclusion"]
for j, h in enumerate(rank_heads, 1):
    ws4.cell(r, j, h)
style_header_row(ws4, r, len(rank_heads))
r += 1

conclusions = [
    "Most critical — removing this collapses risk separation most",
    "Second most critical — significant but recoverable",
    "Least critical — marginal impact on separation",
]
for rank_i, (var, lost, pct) in enumerate(var_ranking):
    vals = [f"#{rank_i+1}", var, f"{lost:.2f}pp", f"{pct:.1f}%", conclusions[rank_i]]
    for j, v in enumerate(vals, 1):
        ws4.cell(r, j, v)
    style_data_row(ws4, r, len(rank_heads), rank_fills[rank_i])
    r += 1

r += 1
write_subheader(ws4, r, "FINAL INSIGHT", 7)
r += 1
final_lines = [
    f"Most critical variable: {var_ranking[0][0]} — removing it loses {var_ranking[0][1]:.2f}pp ({var_ranking[0][2]:.1f}% of total gap).",
    f"Least critical variable: {var_ranking[-1][0]} — removing it only loses {var_ranking[-1][1]:.2f}pp ({var_ranking[-1][2]:.1f}% of total gap).",
    f"This confirms the signal hierarchy found in Task 1: FICO dominates, Term is second, DTI adds marginal separation.",
    f"Business implication: if forced to use only 1 variable → use FICO. If forced to use 2 → use FICO + Term.",
]
for line in final_lines:
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws4.cell(r, 1, line)
    c.font = Font(size=10)
    c.alignment = ALIGN_L
    c.border = thin_border()
    r += 1

autofit(ws4)

# =========================
# SAVE
# =========================
wb.save(OUT_PATH)
print(f"\n✅ DONE → {OUT_PATH}")
print(f"Sheets: Total_Exposure | Biggest_Risk_Analysis | Business_Decisions | Variable_Importance")
