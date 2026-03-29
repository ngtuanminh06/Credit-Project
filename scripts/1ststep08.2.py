import os
import re
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================
# CONFIG
# =========================
FILEPATH = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\accepted_2007_to_2018Q4.csv.gz"

OUT_CLEAN_CSV = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\week6_clean_data.csv"
OUT_XLSX      = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\week6_all_in_one.xlsx"

CHUNKSIZE = 300_000
SEED = 42
SAMPLE_N = 100

USECOLS = [
    "id",
    "loan_status",
    "loan_amnt",
    "funded_amnt",          # ← THÊM: để tính EAD chính xác
    "recoveries",           # ← THÊM: để tính LGD thực tế (Recovery Given Default)
    "fico_range_low",
    "fico_range_high",
    "dti",
    "term"
]

# Risk mapping (industry-like)
LOW_STATUSES  = {"Fully Paid", "Current"}
MED_STATUSES  = {"In Grace Period", "Late (16-30 days)"}
HIGH_STATUSES = {"Late (31-120 days)", "Default", "Charged Off"}

# Excel styles
FILL_HEADER  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_METRIC  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_RED     = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
FILL_YELLOW  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_GREEN   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FONT_BOLD    = Font(bold=True)
ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")

rng = np.random.default_rng(SEED)

# =========================
# Helpers
# =========================
def to_numeric_safe(s):
    return pd.to_numeric(s, errors="coerce")

def clean_term(term_series):
    s = term_series.astype(str).str.strip()
    return pd.to_numeric(s.str.extract(r"(\d+)")[0], errors="coerce")

def normalize_loan_status(x):
    s = x.astype("string").str.strip()
    return s.replace({"nan": pd.NA, "NaN": pd.NA, "None": pd.NA, "": pd.NA})

def calc_fico_avg(df):
    low  = to_numeric_safe(df["fico_range_low"])
    high = to_numeric_safe(df["fico_range_high"])
    return (low + high) / 2.0

def make_risk_label_3class(df):
    ls   = normalize_loan_status(df["loan_status"])
    low  = ls.isin(list(LOW_STATUSES))
    med  = ls.isin(list(MED_STATUSES))
    high = ls.isin(list(HIGH_STATUSES))
    out  = np.select([low, med, high],
                     ["low_risk", "medium_risk", "high_risk"],
                     default="unknown")
    return pd.Series(out, index=df.index, dtype="object")

def write_table(ws, df, start_row, title):
    r = start_row
    ws.cell(r, 1, title).font = Font(bold=True, size=12)
    r += 1
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(r, j, str(col))
        c.font = FONT_BOLD; c.fill = FILL_HEADER; c.alignment = ALIGN_WRAP
    r += 1
    for row in df.itertuples(index=False):
        for j, v in enumerate(row, start=1):
            ws.cell(r, j, v)
        r += 1
    return r + 1

def autofit(ws, max_scan_rows=2000, min_w=10, max_w=55):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells[:max_scan_rows]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)

def group_3level(x, q33, q66, low_label, mid_label, high_label):
    out  = pd.Series(pd.NA, index=x.index, dtype="object")
    mask = x.notna()
    out.loc[mask & (x <= q33)]              = low_label
    out.loc[mask & (x > q33) & (x <= q66)] = mid_label
    out.loc[mask & (x > q66)]              = high_label
    return out

# =========================
# PASS 0: estimate cutpoints
# =========================
def estimate_cutpoints():
    fico_vals, dti_vals = [], []
    max_collect = 300_000
    total = 0

    for chunk in pd.read_csv(FILEPATH, usecols=USECOLS, chunksize=CHUNKSIZE, low_memory=False):
        total += len(chunk)
        chunk["fico_avg"] = calc_fico_avg(chunk)
        chunk["dti"]      = to_numeric_safe(chunk["dti"])

        f = chunk["fico_avg"].dropna().to_numpy()
        d = chunk["dti"].dropna().to_numpy()

        if f.size:
            take = min(20_000, f.size)
            fico_vals.extend(rng.choice(f, size=take, replace=False).tolist())
        if d.size:
            take = min(20_000, d.size)
            dti_vals.extend(rng.choice(d, size=take, replace=False).tolist())

        if len(fico_vals) >= max_collect and len(dti_vals) >= max_collect:
            break
        if total % (CHUNKSIZE * 2) == 0:
            print(f"  scanned {total:,} rows...")

    fico_arr = np.array(fico_vals, dtype=float)
    dti_arr  = np.array(dti_vals,  dtype=float)

    fico_q33 = float(np.quantile(fico_arr, 0.33)) if fico_arr.size else np.nan
    fico_q66 = float(np.quantile(fico_arr, 0.66)) if fico_arr.size else np.nan
    dti_q33  = float(np.quantile(dti_arr,  0.33)) if dti_arr.size  else np.nan
    dti_q66  = float(np.quantile(dti_arr,  0.66)) if dti_arr.size  else np.nan

    return fico_q33, fico_q66, dti_q33, dti_q66

# =========================
# MAIN
# =========================
def main():
    print("=" * 90)
    print("WEEK 6 v2 — CLEAN_DATA CSV + EXCEL (OVERVIEW + SAMPLE_100 + NULL_CHECK)")
    print("=" * 90)
    print(f"Input : {FILEPATH}")
    print(f"CSV   : {OUT_CLEAN_CSV}")
    print(f"Excel : {OUT_XLSX}")
    print(f"Chunk : {CHUNKSIZE:,}\n")

    print("PASS 0 — estimating cutpoints for grouping (fico_avg + dti)...")
    fico_q33, fico_q66, dti_q33, dti_q66 = estimate_cutpoints()
    print(f"fico_avg cutpoints: q33={fico_q33:.2f}, q66={fico_q66:.2f}")
    print(f"dti     cutpoints: q33={dti_q33:.2f},  q66={dti_q66:.2f}\n")

    if os.path.exists(OUT_CLEAN_CSV):
        os.remove(OUT_CLEAN_CSV)

    # Accumulators
    total_rows         = 0
    usable_rows        = 0
    loan_status_counts = {}
    risk_counts        = {"low_risk": 0, "medium_risk": 0, "high_risk": 0, "unknown": 0}
    grp_counts         = {"fico_group": {}, "dti_group": {}, "term_group": {}}

    null_raw     = {col: 0 for col in USECOLS}
    null_derived = {"fico_avg": 0, "term_m": 0, "loan_amnt_numeric": 0}

    sample_rows    = []
    seen_usable    = 0
    header_written = False
    processed      = 0

    print("PASS 1 — writing CLEAN_DATA CSV + building OVERVIEW + SAMPLE_100 + NULL_CHECK...")

    for chunk in pd.read_csv(FILEPATH, usecols=USECOLS, chunksize=CHUNKSIZE, low_memory=False):
        processed  += len(chunk)
        total_rows += len(chunk)

        # Đếm null raw
        for col in USECOLS:
            if col in chunk.columns:
                null_raw[col] += int(chunk[col].isna().sum())

        # Feature engineering
        chunk["loan_status"]  = normalize_loan_status(chunk["loan_status"])
        chunk["fico_avg"]     = calc_fico_avg(chunk)
        chunk["dti"]          = to_numeric_safe(chunk["dti"])
        chunk["term_m"]       = clean_term(chunk["term"])
        chunk["loan_amnt"]    = to_numeric_safe(chunk["loan_amnt"])
        chunk["funded_amnt"]  = to_numeric_safe(chunk["funded_amnt"])   # ← THÊM
        chunk["recoveries"]   = to_numeric_safe(chunk["recoveries"])    # ← THÊM
        chunk["risk_label"]   = make_risk_label_3class(chunk)

        # Đếm null derived
        null_derived["fico_avg"]         += int(chunk["fico_avg"].isna().sum())
        null_derived["term_m"]           += int(chunk["term_m"].isna().sum())
        null_derived["loan_amnt_numeric"]+= int(chunk["loan_amnt"].isna().sum())

        # Filter usable
        usable = chunk.dropna(subset=["fico_avg", "dti", "term_m"]).copy()
        usable = usable[usable["risk_label"] != "unknown"].copy()
        usable_rows += len(usable)

        # Grouping
        usable["fico_group"] = group_3level(usable["fico_avg"], fico_q33, fico_q66,
                                            "low_fico", "mid_fico", "high_fico")
        usable["dti_group"]  = group_3level(usable["dti"], dti_q33, dti_q66,
                                            "low_dti", "mid_dti", "high_dti")
        usable["term_group"] = np.where(usable["term_m"] == 60, "term_60", "term_36")

        # Loan status counts
        vc_ls = chunk["loan_status"].value_counts(dropna=False)
        for k, v in vc_ls.items():
            loan_status_counts[str(k)] = loan_status_counts.get(str(k), 0) + int(v)

        # Risk counts
        vc_r = usable["risk_label"].value_counts(dropna=False)
        for k, v in vc_r.items():
            key = str(k)
            if key in risk_counts:
                risk_counts[key] += int(v)

        # Group counts
        for grp_col in ["fico_group", "dti_group", "term_group"]:
            tmp   = usable.groupby([grp_col, "risk_label"], dropna=False).size()
            store = grp_counts[grp_col]
            for (g, r2), n in tmp.items():
                gk = str(g); rk = str(r2)
                if gk not in store:
                    store[gk] = {"low_risk": 0, "medium_risk": 0, "high_risk": 0}
                store[gk][rk] = store[gk].get(rk, 0) + int(n)

        # out_cols — bao gồm funded_amnt và recoveries
        out_cols = [
            "id", "loan_status",
            "loan_amnt",
            "funded_amnt",                          # ← THÊM
            "recoveries",                           # ← THÊM
            "fico_range_low", "fico_range_high",
            "fico_avg", "fico_group",
            "dti", "dti_group",
            "term", "term_m", "term_group",
            "risk_label"
        ]
        usable[out_cols].to_csv(
            OUT_CLEAN_CSV, mode="a", index=False, header=(not header_written)
        )
        header_written = True

        # Reservoir sampling
        for _, row in usable[out_cols].iterrows():
            seen_usable += 1
            if len(sample_rows) < SAMPLE_N:
                sample_rows.append(row.to_dict())
            else:
                j = rng.integers(0, seen_usable)
                if j < SAMPLE_N:
                    sample_rows[j] = row.to_dict()

        if processed % (CHUNKSIZE * 2) == 0:
            print(f"  progress {processed:,} / ? | written={usable_rows:,} | sampled_from={seen_usable:,}", end="\r")

    print("\nPASS 1 done.")
    print(f"Total rows in file : {total_rows:,}")
    print(f"Rows usable (CSV)  : {usable_rows:,}")

    # Print null summary
    print(f"\nNULL SUMMARY (tính trên {total_rows:,} rows):")
    print(f"  {'Column':<28} {'Null':>10}  {'% Null':>7}")
    print(f"  {'-'*50}")
    for col in USECOLS:
        cnt = null_raw[col]
        pct = cnt / total_rows * 100
        print(f"  [raw]     {col:<20} {cnt:>10,}  {pct:>6.2f}%")
    for col, cnt in null_derived.items():
        pct = cnt / total_rows * 100
        print(f"  [derived] {col:<20} {cnt:>10,}  {pct:>6.2f}%")

    # =========================
    # BUILD TABLES
    # =========================
    ls_df = (
        pd.DataFrame({"loan_status": list(loan_status_counts.keys()),
                      "count": list(loan_status_counts.values())})
        .sort_values("count", ascending=False).reset_index(drop=True)
    )
    ls_df["pct"] = (ls_df["count"] / total_rows * 100).round(4)

    risk_df = pd.DataFrame({
        "risk_label": ["low_risk", "medium_risk", "high_risk"],
        "count": [risk_counts["low_risk"], risk_counts["medium_risk"], risk_counts["high_risk"]]
    })
    risk_df["pct"] = (risk_df["count"] / max(usable_rows, 1) * 100).round(4)

    def build_group_summary(grp_col, feature_name):
        rows  = []
        store = grp_counts[grp_col]
        for gk, d in store.items():
            n = d.get("low_risk", 0) + d.get("medium_risk", 0) + d.get("high_risk", 0)
            if n == 0: continue
            rows.append({
                "feature":         feature_name,
                "group":           gk,
                "n":               n,
                "low_risk_pct":    round(d.get("low_risk",    0) / n * 100, 2),
                "medium_risk_pct": round(d.get("medium_risk", 0) / n * 100, 2),
                "high_risk_pct":   round(d.get("high_risk",   0) / n * 100, 2),
            })
        return pd.DataFrame(rows).sort_values(["feature", "high_risk_pct"], ascending=[True, False])

    group_df = pd.concat([
        build_group_summary("fico_group", "fico_avg"),
        build_group_summary("dti_group",  "dti"),
        build_group_summary("term_group", "term"),
    ], ignore_index=True)

    sample_df = pd.DataFrame(sample_rows)

    # NULL CHECK dataframe
    null_excel_rows = []
    for col in USECOLS:
        cnt = null_raw[col]
        pct = round(cnt / total_rows * 100, 4)
        null_excel_rows.append({
            "Type":       "raw",
            "Column":     col,
            "Null Count": cnt,
            "Total Rows": total_rows,
            "% Null":     pct,
            "Status":     ("CRITICAL (>20%)" if pct > 20
                           else "WARNING (5-20%)" if pct > 5
                           else "LOW (<5%)"        if pct > 0
                           else "OK (no nulls)"),
        })
    for col, cnt in null_derived.items():
        pct = round(cnt / total_rows * 100, 4)
        null_excel_rows.append({
            "Type":       "derived",
            "Column":     col,
            "Null Count": cnt,
            "Total Rows": total_rows,
            "% Null":     pct,
            "Status":     ("CRITICAL (>20%)" if pct > 20
                           else "WARNING (5-20%)" if pct > 5
                           else "LOW (<5%)"        if pct > 0
                           else "OK (no nulls)"),
        })

    null_df = (pd.DataFrame(null_excel_rows)
               .sort_values("% Null", ascending=False)
               .reset_index(drop=True))

    # =========================
    # WRITE EXCEL
    # =========================
    wb = Workbook()

    # Sheet 1: OVERVIEW
    ws1 = wb.active
    ws1.title = "OVERVIEW"
    r = 1
    r = write_table(ws1, ls_df.head(20), r, "Loan status distribution (top 20)")
    r = write_table(ws1, risk_df,         r, "Risk label distribution (within usable rows)")
    r = write_table(ws1, group_df,         r, "Group summary: % risk by 3 groups (fico_avg, dti, term)")
    autofit(ws1)

    # Sheet 2: SAMPLE_100
    ws2 = wb.create_sheet("SAMPLE_100")
    for j, col in enumerate(sample_df.columns, start=1):
        c = ws2.cell(1, j, str(col))
        c.font = FONT_BOLD; c.fill = FILL_HEADER; c.alignment = ALIGN_WRAP
    for i, row in enumerate(sample_df.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            ws2.cell(i, j, v)

    metric_cols = {"fico_avg", "fico_group", "dti_group", "term_m",
                   "term_group", "risk_label", "loan_amnt",
                   "funded_amnt", "recoveries"}         # ← THÊM 2 cột mới vào highlight
    col_to_idx = {c: idx + 1 for idx, c in enumerate(sample_df.columns)}
    for c in metric_cols:
        if c in col_to_idx:
            j = col_to_idx[c]
            for i in range(1, len(sample_df) + 2):
                ws2.cell(i, j).fill = FILL_METRIC
                if i == 1:
                    ws2.cell(i, j).font = FONT_BOLD
    autofit(ws2)

    # Sheet 3: NULL_CHECK
    ws3 = wb.create_sheet("NULL_CHECK")
    ws3.cell(1, 1, f"NULL CHECK — tổng {total_rows:,} rows").font = Font(bold=True, size=12)
    ws3.cell(2, 1, "raw = cột đọc thẳng từ CSV  |  derived = cột tính sau engineering").font = Font(italic=True, color="595959")

    null_headers = ["Type", "Column", "Null Count", "Total Rows", "% Null", "Status"]
    for j, h in enumerate(null_headers, start=1):
        c = ws3.cell(4, j, h)
        c.font = FONT_BOLD; c.fill = FILL_HEADER; c.alignment = ALIGN_WRAP

    for i, row in enumerate(null_df.itertuples(index=False), start=5):
        status = row.Status
        pct    = row._4
        if   "CRITICAL" in status: f = FILL_RED
        elif "WARNING"  in status: f = FILL_YELLOW
        elif "OK"       in status: f = FILL_GREEN
        else:                      f = None

        vals = [row.Type, row.Column,
                f"{row._2:,}", f"{row._3:,}",
                f"{pct:.4f}%", status]
        for j, v in enumerate(vals, start=1):
            c = ws3.cell(i, j, v)
            if f: c.fill = f

    autofit(ws3)

    out_dir = os.path.dirname(OUT_XLSX)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(OUT_XLSX)

    print(f"\n✅ DONE")
    print(f"Saved CSV  : {OUT_CLEAN_CSV}  (now includes funded_amnt + recoveries)")
    print(f"Saved Excel: {OUT_XLSX}")
    print(f"Sheets     : OVERVIEW | SAMPLE_100 | NULL_CHECK")

if __name__ == "__main__":
    main()
