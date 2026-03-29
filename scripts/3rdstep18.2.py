import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
INPUT_CLEAN  = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\week6_clean_data.csv"
OUTPUT_XLSX  = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\step2_5_segmentation_v5.xlsx"

FICO_BUCKET_WIDTH = 20
FICO_THRESHOLD_OVERRIDE = None   # auto-detect → 680

# =========================
# STYLES
# =========================
def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

FILL_HEADER = make_fill("1F4E79")
FILL_SUB    = make_fill("2E75B6")
FILL_GREEN  = make_fill("70AD47")
FILL_WARN   = make_fill("FFD966")
FILL_RED    = make_fill("FF4D4D")
FILL_LITE_RED   = make_fill("FF9999")
FILL_LITE_GREEN = make_fill("C6EFCE")
FILL_LITE_YEL   = make_fill("FFEB9C")
FILL_LITE_ORA   = make_fill("FFCC99")
FILL_SEG = [
    make_fill("C6EFCE"),
    make_fill("FFEB9C"),
    make_fill("FFCC99"),
    make_fill("FF9999"),
]
FILL_GREY = make_fill("F2F2F2")

FONT_WB     = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_ITALIC = Font(italic=True, size=10, color="595959")
FONT_SMALL  = Font(italic=True, size=9,  color="404040")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, ncols, title, fill=FILL_HEADER):
    ws.row_dimensions[row].height = 28
    for j in range(1, ncols + 1):
        c = ws.cell(row, j)
        c.fill = fill; c.font = FONT_WB
        c.alignment = ALIGN_C; c.border = thin_border()
    ws.cell(row, 1, title)

def style_col_headers(ws, row, headers, fill=FILL_SUB):
    ws.row_dimensions[row].height = 22
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.fill = fill; c.font = FONT_WB
        c.alignment = ALIGN_C; c.border = thin_border()

def write_data_row(ws, row, values, fill=None, font=None, height=18):
    ws.row_dimensions[row].height = height
    for j, v in enumerate(values, 1):
        c = ws.cell(row, j, v)
        if fill: c.fill = fill
        c.font = font or FONT_NORMAL
        c.alignment = ALIGN_C; c.border = thin_border()

def autofit(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        try:
            ltr = col_cells[0].column_letter
        except AttributeError:
            continue
        mx = max((len(str(c.value or "")) for c in col_cells if c.value), default=8)
        ws.column_dimensions[ltr].width = min(max(min_w, mx + 2), max_w)

# =========================
# LOAD DATA
# =========================
print("=" * 65)
print("STEP 2-5 v5: SEGMENTATION BY FICO + TERM  (Task 2 & 3 added)")
print("=" * 65)

df = pd.read_csv(INPUT_CLEAN)
print(f"  Rows loaded : {len(df):,}")

df["fico_avg"] = pd.to_numeric(df["fico_avg"], errors="coerce")
df["dti"]      = pd.to_numeric(df["dti"],      errors="coerce")
df["term_m"]   = pd.to_numeric(df["term_m"],   errors="coerce")

df = df.dropna(subset=["fico_avg", "dti", "term_m", "risk_label"])
df = df[df["risk_label"] != "unknown"].copy()

is_high      = df["risk_label"] == "high_risk"
total_rows   = len(df)
baseline_pct = is_high.mean() * 100

print(f"  Rows after filter : {total_rows:,}")
print(f"  Baseline high_risk: {baseline_pct:.2f}%")

# =========================
# BƯỚC 2A — FICO BUCKET ANALYSIS
# =========================
def bucket_analysis(series, is_high_mask, bucket_width, min_count=200):
    s_ref  = series.dropna()
    lo     = (s_ref.min() // bucket_width) * bucket_width
    hi     = (s_ref.max() // bucket_width + 1) * bucket_width
    edges  = np.arange(lo, hi + bucket_width, bucket_width)
    rows   = []
    for i in range(len(edges) - 1):
        left, right = edges[i], edges[i + 1]
        mask = (series >= left) & (series < right)
        n    = int(mask.sum())
        if n < min_count:
            continue
        pct_high = is_high_mask[mask].mean() * 100
        rows.append({
            "bucket":            f"{int(left)}-{int(right)}",
            "bucket_lo":         left,
            "n":                 n,
            "pct_total":         round(n / total_rows * 100, 2),
            "high_risk_pct":     round(pct_high, 2),
            "delta_vs_baseline": round(pct_high - baseline_pct, 2),
        })
    bdf = pd.DataFrame(rows).sort_values("bucket_lo").reset_index(drop=True)
    threshold_bucket, jump_size = None, 0.0
    if len(bdf) >= 2:
        diffs   = bdf["high_risk_pct"].diff().abs()
        idx_max = diffs.idxmax()
        if not pd.isna(idx_max):
            threshold_bucket = float(bdf.loc[idx_max, "bucket_lo"])
            jump_size        = float(diffs[idx_max])
    return bdf, threshold_bucket, jump_size

print("\n--- Bước 2A: FICO threshold ---")
fico_bdf, fico_auto, fico_jump = bucket_analysis(df["fico_avg"], is_high, FICO_BUCKET_WIDTH)
fico_cut = FICO_THRESHOLD_OVERRIDE if FICO_THRESHOLD_OVERRIDE is not None else fico_auto
print(f"  FICO cut = {fico_cut}  (jump {fico_jump:.2f}pp)")

# =========================
# BƯỚC 2B — TERM ANALYSIS
# =========================
print("\n--- Bước 2B: Term analysis ---")
term_rows = []
for tm, label in [(36, "term_36 (3 năm)"), (60, "term_60 (5 năm)")]:
    mask     = df["term_m"] == tm
    n        = int(mask.sum())
    pct_high = is_high[mask].mean() * 100
    term_rows.append({
        "term": label, "term_m": tm, "n": n,
        "pct_total":         round(n / total_rows * 100, 2),
        "high_risk_pct":     round(pct_high, 2),
        "delta_vs_baseline": round(pct_high - baseline_pct, 2),
        "lift":              round(pct_high / baseline_pct, 2),
    })
    print(f"  {label:20s} | n={n:>8,} | high_risk={pct_high:.2f}% | delta={pct_high-baseline_pct:+.2f}pp")

term_df  = pd.DataFrame(term_rows)
term_gap = term_rows[1]["high_risk_pct"] - term_rows[0]["high_risk_pct"]

# =========================
# BƯỚC 3 — GÁN SEGMENT
# =========================
def assign_segment(row):
    is_prime = row["fico_avg"] >= fico_cut
    is_36m   = row["term_m"]   == 36
    if   is_prime and is_36m:     return 1
    elif is_prime and not is_36m: return 2
    elif not is_prime and is_36m: return 3
    else:                          return 4

df["segment"] = df.apply(assign_segment, axis=1)

SEG_NAMES = {1: "Prime Short-Term", 2: "Prime Long-Term",
             3: "Subprime Short-Term", 4: "Subprime Long-Term"}
SEG_DEF   = {1: f"FICO ≥ {fico_cut} + Term 36M",
             2: f"FICO ≥ {fico_cut} + Term 60M",
             3: f"FICO < {fico_cut} + Term 36M",
             4: f"FICO < {fico_cut} + Term 60M"}

# =========================
# BƯỚC 4 — ĐỊNH LƯỢNG
# =========================
print("\n--- Bước 4: Quantifying ---")
seg_rows = []
for sid in [1, 2, 3, 4]:
    sub      = df[df["segment"] == sid]
    n        = len(sub)
    pct_tot  = n / total_rows * 100
    pct_high = sub["risk_label"].eq("high_risk").mean()   * 100
    pct_med  = sub["risk_label"].eq("medium_risk").mean() * 100
    pct_low  = sub["risk_label"].eq("low_risk").mean()    * 100
    avg_fico = sub["fico_avg"].mean()
    avg_dti  = sub["dti"].mean()
    lift     = pct_high / baseline_pct
    avg_loan = sub["loan_amnt"].mean() if "loan_amnt" in sub.columns else None

    seg_rows.append({
        "segment_id": sid, "segment_name": SEG_NAMES[sid],
        "definition": SEG_DEF[sid],
        "n": n, "pct_of_total": round(pct_tot, 2),
        "avg_fico": round(avg_fico, 1), "avg_dti": round(avg_dti, 2),
        "avg_loan_amnt": round(avg_loan, 0) if avg_loan else "N/A",
        "high_risk_pct": round(pct_high, 2), "medium_risk_pct": round(pct_med, 2),
        "low_risk_pct":  round(pct_low,  2), "lift": round(lift, 2),
        "delta_pp": round(pct_high - baseline_pct, 2),
    })
    print(f"  Seg {sid} | {SEG_NAMES[sid]:24s} | n={n:>8,} ({pct_tot:5.1f}%) "
          f"| high_risk={pct_high:5.2f}% | lift={lift:.2f}x")

seg_df = pd.DataFrame(seg_rows)

# =========================
# BƯỚC 5 — SO WHAT
# =========================
def build_analysis(row):
    sid   = row["segment_id"]
    n     = row["n"];  pct_t = row["pct_of_total"]
    pct_h = row["high_risk_pct"]; lift = row["lift"]; d_pp = row["delta_pp"]
    avg_f = row["avg_fico"]; avg_d = row["avg_dti"]
    sign  = "cao hơn" if d_pp > 0 else "thấp hơn"
    data  = {
        1: (
            f"'Prime Short-Term' gồm {n:,} khách hàng ({pct_t:.1f}% danh mục). "
            f"FICO tốt (avg {avg_f:.0f}), chọn kỳ hạn 36 tháng — hành vi tài chính tự tin nhất. "
            f"DTI trung bình {avg_d:.1f}% (mô tả).",
            f"High_risk chỉ {pct_h:.1f}%, thấp hơn baseline {abs(d_pp):.1f}pp (lift={lift:.2f}x). "
            f"Kỳ hạn ngắn giảm thời gian phơi rủi ro, FICO tốt đảm bảo lịch sử trả nợ.",
            "Ưu tiên tăng trưởng. Fast-track approval và ưu đãi lãi suất để giữ chân nhóm này."
        ),
        2: (
            f"'Prime Long-Term' gồm {n:,} khách hàng ({pct_t:.1f}% danh mục). "
            f"FICO tốt (avg {avg_f:.0f}) nhưng cần 5 năm để trả — áp lực tài chính rõ hơn. "
            f"DTI trung bình {avg_d:.1f}%.",
            f"High_risk {pct_h:.1f}%, {sign} baseline {abs(d_pp):.1f}pp (lift={lift:.2f}x). "
            f"Kỳ hạn dài kéo dài thời gian phơi rủi ro — 5 năm đủ để thu nhập biến động.",
            "Theo dõi sát thu nhập và DTI. Khuyến khích refinance sang 36 tháng nếu đủ điều kiện."
        ),
        3: (
            f"'Subprime Short-Term' gồm {n:,} khách hàng ({pct_t:.1f}% danh mục). "
            f"FICO dưới chuẩn (avg {avg_f:.0f}) nhưng chọn trả nhanh — tự giới hạn rủi ro. "
            f"DTI trung bình {avg_d:.1f}%.",
            f"High_risk {pct_h:.1f}%, {sign} baseline {abs(d_pp):.1f}pp (lift={lift:.2f}x). "
            f"Kỳ hạn ngắn bù đắp một phần cho FICO yếu. Rủi ro tương đương Seg 2 dù đặc điểm khác.",
            "Không từ chối hoàn toàn. Giới hạn số tiền vay, giữ kỳ hạn 36 tháng là điều kiện bắt buộc."
        ),
        4: (
            f"'Subprime Long-Term' gồm {n:,} khách hàng ({pct_t:.1f}% danh mục). "
            f"Nhóm nguy hiểm nhất: FICO yếu (avg {avg_f:.0f}) + kỳ hạn 60 tháng. "
            f"DTI trung bình {avg_d:.1f}%.",
            f"High_risk {pct_h:.1f}%, {sign} baseline {abs(d_pp):.1f}pp (lift={lift:.2f}x). "
            f"FICO yếu + 5 năm phơi rủi ro = tổ hợp nguy hiểm nhất. Cứ 4 người thì 1 người vỡ nợ.",
            "Kiểm soát nghiêm nhất: giảm hạn mức vay, không khuyến khích 60 tháng, "
            "áp lãi suất bù rủi ro. Early warning hàng quý."
        ),
    }
    return data[sid]

# =========================
# TASK 2 — SUMMARY SENTENCES CHO MỖI BOUNDARY
# =========================
# Được tính từ fico_bdf và term_df
# Viết sẵn vào Excel như 1 bảng diễn giải

def get_fico_summary_sentences(bdf, baseline):
    sentences = []
    for i in range(1, len(bdf)):
        prev = bdf.iloc[i-1]
        curr = bdf.iloc[i]
        jump = curr["high_risk_pct"] - prev["high_risk_pct"]
        is_threshold = (curr["bucket_lo"] == fico_cut)
        if is_threshold:
            note = (f"⚠️ NGƯỠNG CẮT: Vượt FICO {int(curr['bucket_lo'])}, rủi ro giảm {abs(jump):.1f}pp "
                    f"(từ {prev['high_risk_pct']:.1f}% → {curr['high_risk_pct']:.1f}%). "
                    f"Đây là ranh giới lớn nhất trong toàn thang FICO.")
        elif abs(jump) >= 2:
            direction = "giảm" if jump < 0 else "tăng"
            note = (f"Rủi ro {direction} {abs(jump):.1f}pp "
                    f"({prev['high_risk_pct']:.1f}% → {curr['high_risk_pct']:.1f}%). "
                    f"Mỗi 20 điểm FICO {'cao hơn' if jump<0 else 'thấp hơn'} tương ứng giảm/tăng rủi ro đáng kể.")
        else:
            note = (f"Rủi ro thay đổi nhỏ {jump:+.1f}pp "
                    f"({prev['high_risk_pct']:.1f}% → {curr['high_risk_pct']:.1f}%). "
                    f"Vùng FICO này có rủi ro ổn định, thay đổi dần đều.")
        sentences.append({
            "Boundary":        f"{prev['bucket']} → {curr['bucket']}",
            "Δ High Risk (pp)": f"{jump:+.2f}",
            "Pattern":         "Jump" if abs(jump) >= 3 else "Smooth",
            "Summary":         note,
            "⚠ Threshold?":    "✅ THRESHOLD" if is_threshold else "",
        })
    return sentences

# =========================
# TASK 3 — SEGMENT HYPOTHESIS (trước khi nhìn data)
# =========================
task3_data = [
    {
        "Segment":        "Segment 1 — Prime Short-Term",
        "Conditions":     f"FICO ≥ {fico_cut}  +  Term = 36 tháng",
        "Expected Risk":  "Low (< 10%)",
        "Lý do chọn":     (
            "FICO tốt = lịch sử trả nợ đáng tin cậy. "
            "Term 36M = cam kết trả nhanh, ít thời gian để hoàn cảnh thay đổi xấu. "
            "Kết hợp 2 yếu tố tốt → kỳ vọng rủi ro thấp nhất."
        ),
        "Predicted lift": "< 0.80x",
    },
    {
        "Segment":        "Segment 2 — Prime Long-Term",
        "Conditions":     f"FICO ≥ {fico_cut}  +  Term = 60 tháng",
        "Expected Risk":  "Moderate (12–16%)",
        "Lý do chọn":     (
            "FICO tốt nhưng cần 5 năm để trả → áp lực tài chính cao hơn. "
            "Kỳ hạn dài = nhiều thời gian để thu nhập hoặc kinh tế thay đổi. "
            "Rủi ro 'tiềm ẩn': trông ổn hôm nay nhưng dễ tổn thương dài hạn."
        ),
        "Predicted lift": "1.00–1.20x",
    },
    {
        "Segment":        "Segment 3 — Subprime Short-Term",
        "Conditions":     f"FICO < {fico_cut}  +  Term = 36 tháng",
        "Expected Risk":  "Moderate (14–18%)",
        "Lý do chọn":     (
            "FICO yếu = lịch sử tín dụng có vấn đề. "
            "Nhưng Term 36M = tự giới hạn, muốn trả nhanh → bù đắp một phần. "
            "Kỳ vọng rủi ro trung bình, không nguy hiểm nhất."
        ),
        "Predicted lift": "1.10–1.30x",
    },
    {
        "Segment":        "Segment 4 — Subprime Long-Term",
        "Conditions":     f"FICO < {fico_cut}  +  Term = 60 tháng",
        "Expected Risk":  "High (> 20%)",
        "Lý do chọn":     (
            "FICO yếu + Term dài = 2 yếu tố xấu cộng lại. "
            "Lịch sử tín dụng kém + 5 năm phơi rủi ro = tổ hợp nguy hiểm nhất. "
            "Kỳ vọng rủi ro cao nhất trong 4 nhóm."
        ),
        "Predicted lift": "> 1.50x",
    },
]

# =========================
# WRITE EXCEL
# =========================
print("\nWriting Excel ...")
wb = Workbook()

# ── SHEET 1: Threshold FICO (Task 2) ──
ws1 = wb.active
ws1.title = "Task2_FICO_Threshold"
ws1.freeze_panes = "A4"
r = 1
style_header(ws1, r, 6,
    f"TASK 2 — THRESHOLD FICO  (bucket={FICO_BUCKET_WIDTH}pts | baseline={baseline_pct:.2f}%)")
r += 1
ws1.cell(r, 1,
    "Câu hỏi: Rủi ro tăng đều đặn hay nhảy đột ngột? Ngưỡng nào là ranh giới rủi ro thực sự?"
).font = FONT_ITALIC
ws1.row_dimensions[r].height = 18
r += 2

style_col_headers(ws1, r,
    ["FICO Bucket", "N (loans)", "% of Total", "High Risk %", "Δ vs Baseline (pp)", "⚠ Ghi chú"])
r += 1

for _, row in fico_bdf.iterrows():
    is_t  = (row["bucket_lo"] == fico_cut)
    fill  = FILL_GREEN if is_t else (FILL_WARN if row["high_risk_pct"] > baseline_pct * 1.3 else None)
    flag  = f"✅ THRESHOLD (jump +{fico_jump:.1f}pp) — ranh giới prime vs subprime" if is_t else ""
    write_data_row(ws1, r, [
        row["bucket"], f"{row['n']:,}", f"{row['pct_total']:.2f}%",
        f"{row['high_risk_pct']:.2f}%", f"{row['delta_vs_baseline']:+.2f}", flag,
    ], fill=fill, font=FONT_BOLD if is_t else FONT_NORMAL)
    r += 1

# Summary sentences (Task 2)
r += 1
ws1.cell(r, 1, "SUMMARY SENTENCES — MỖI BOUNDARY").font = FONT_BOLD
r += 1
style_col_headers(ws1, r,
    ["Boundary", "Δ High Risk (pp)", "Pattern", "Summary sentence", "⚠ Threshold?"])
r += 1

summaries = get_fico_summary_sentences(fico_bdf, baseline_pct)
for s in summaries:
    is_t = s["⚠ Threshold?"] != ""
    fill = FILL_GREEN if is_t else (FILL_WARN if s["Pattern"] == "Jump" else None)
    write_data_row(ws1, r, [
        s["Boundary"], s["Δ High Risk (pp)"], s["Pattern"],
        s["Summary"], s["⚠ Threshold?"],
    ], fill=fill, font=FONT_BOLD if is_t else FONT_NORMAL, height=40)
    r += 1

r += 1
ws1.cell(r, 1,
    "Ghi chú: LendingClub yêu cầu FICO ≥ 660 → không có bucket nào dưới 660. "
    "Câu hỏi 'Does risk spike below 600?' không thể trả lời từ dataset này."
).font = FONT_ITALIC
ws1.row_dimensions[r].height = 30

autofit(ws1)
ws1.column_dimensions["D"].width = 70
ws1.column_dimensions["E"].width = 50

# ── SHEET 2: Term Analysis (Task 2) ──
ws2 = wb.create_sheet("Task2_Term_Analysis")
r = 1
style_header(ws2, r, 7,
    f"TASK 2 — TERM ANALYSIS  (binary: 36M vs 60M | baseline={baseline_pct:.2f}%)")
r += 1
ws2.cell(r, 1,
    "Term đã là binary — không cần bucket. So sánh trực tiếp 36M vs 60M."
).font = FONT_ITALIC
ws2.row_dimensions[r].height = 18
r += 2

style_col_headers(ws2, r,
    ["Kỳ hạn", "N (loans)", "% of Total", "High Risk %", "Δ vs Baseline (pp)", "Lift (x)", "Pattern"])
r += 1

for _, row in term_df.iterrows():
    is_60 = (row["term_m"] == 60)
    fill  = FILL_LITE_RED if is_60 else FILL_LITE_GREEN
    pat   = f"⚠️ HIGH: +{term_gap:.1f}pp vs 36M" if is_60 else "✅ LOW: ranh giới an toàn hơn"
    write_data_row(ws2, r, [
        row["term"], f"{row['n']:,}", f"{row['pct_total']:.2f}%",
        f"{row['high_risk_pct']:.2f}%", f"{row['delta_vs_baseline']:+.2f}",
        f"{row['lift']:.2f}x", pat,
    ], fill=fill, font=FONT_BOLD)
    r += 1

r += 2
ws2.cell(r, 1, "SUMMARY SENTENCES — TERM BOUNDARY").font = FONT_BOLD
r += 1
for note in [
    f"Term 36M → 60M: rủi ro tăng {term_gap:.2f}pp — đây là jump rõ ràng, không phải thay đổi dần đều.",
    f"Term là biến binary nên ranh giới 36M/60M chính là threshold tự nhiên — không cần tìm ngưỡng.",
    f"Người chọn 60M chấp nhận trả nợ trong 5 năm → nhiều thời gian hơn để gặp sự cố tài chính.",
    f"Gap {term_gap:.2f}pp > gap DTI (6.0pp) → Term là tín hiệu mạnh thứ 2 sau FICO.",
]:
    c = ws2.cell(r, 1, note)
    c.font = FONT_SMALL
    c.alignment = ALIGN_L
    ws2.row_dimensions[r].height = 22
    r += 1

autofit(ws2)

# ── SHEET 3: Task 3 — Segment Hypothesis ──
ws3 = wb.create_sheet("Task3_Segment_Hypothesis")
r = 1
style_header(ws3, r, 5,
    "TASK 3 — ĐỊNH NGHĨA SEGMENT (TRƯỚC KHI NHÌN DATA)")
r += 1
ws3.cell(r, 1,
    "Dựa trên logic kinh tế và kết quả bucket analysis — định nghĩa 4 nhóm trước khi quantify."
).font = FONT_ITALIC
ws3.row_dimensions[r].height = 18
r += 2

style_col_headers(ws3, r,
    ["Segment", "Điều kiện (Conditions)", "Expected Risk Level", "Predicted Lift", "Lý do chọn"])
r += 1

for i, seg in enumerate(task3_data):
    fill = FILL_SEG[i]
    write_data_row(ws3, r, [
        seg["Segment"], seg["Conditions"],
        seg["Expected Risk"], seg["Predicted lift"], seg["Lý do chọn"],
    ], fill=fill, font=FONT_BOLD if i in [0, 3] else FONT_NORMAL, height=55)
    r += 1

r += 2
ws3.cell(r, 1, "→ Sheet 'Segments_Table' là kết quả thực tế — so sánh với hypothesis ở trên.").font = FONT_ITALIC
autofit(ws3)
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 65

# ── SHEET 4: Segments Table (Task 4) ──
ws4 = wb.create_sheet("Segments_Table")
ws4.freeze_panes = "A4"
r = 1
ncols = 12
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
style_header(ws4, r, ncols,
    "TASK 4 — QUANTIFY SEGMENTS: FICO × TERM  (DTI = cột mô tả)")
r += 1
ws4.cell(r, 1,
    f"Segmentation: FICO cut = {fico_cut} × Term (36M/60M)  |  "
    f"Baseline = {baseline_pct:.2f}%  |  DTI = mô tả, không dùng để chia"
).font = FONT_ITALIC
r += 1
style_col_headers(ws4, r, [
    "Segment", "Name", "Definition",
    "N (loans)", "% of Total",
    "Avg FICO", "Avg DTI *", "Avg Loan Amt",
    "High Risk %", "Medium Risk %",
    "Δ vs Baseline (pp)", "Lift (x)",
])
r += 1

for row in seg_rows:
    sid = row["segment_id"]
    write_data_row(ws4, r, [
        f"Segment {sid}", row["segment_name"], row["definition"],
        f"{row['n']:,}", f"{row['pct_of_total']:.2f}%",
        f"{row['avg_fico']:.1f}", f"{row['avg_dti']:.1f}",
        f"{row['avg_loan_amnt']:,}" if isinstance(row['avg_loan_amnt'], (int,float)) else row['avg_loan_amnt'],
        f"{row['high_risk_pct']:.2f}%", f"{row['medium_risk_pct']:.2f}%",
        f"{row['delta_pp']:+.2f}", f"{row['lift']:.2f}x",
    ], fill=FILL_SEG[sid-1], font=FONT_BOLD)
    r += 1

r += 1
ws4.cell(r, 1,
    "* Avg DTI: cột mô tả — không dùng để chia segment. "
    "Cho thấy gánh nặng nợ trung bình của từng nhóm sau khi chia theo FICO + Term."
).font = FONT_ITALIC
r += 2

# Risk breakdown sub-table
ws4.cell(r, 1, "Risk Label Breakdown per Segment").font = FONT_BOLD
r += 1
style_col_headers(ws4, r,
    ["Segment", "Name", "Low Risk %", "Medium Risk %", "High Risk %", "Total N"])
r += 1
for row in seg_rows:
    sid = row["segment_id"]
    write_data_row(ws4, r, [
        f"Segment {sid}", row["segment_name"],
        f"{row['low_risk_pct']:.2f}%", f"{row['medium_risk_pct']:.2f}%",
        f"{row['high_risk_pct']:.2f}%", f"{row['n']:,}",
    ], fill=FILL_SEG[sid-1])
    r += 1

autofit(ws4)
ws4.column_dimensions["C"].width = 28
ws4.column_dimensions["G"].width = 14

# ── SHEET 5: Analysis So What (Task 5) ──
ws5 = wb.create_sheet("Task5_Analysis")
r = 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
style_header(ws5, r, 3, "TASK 5 — WRITTEN ANALYSIS: WHO / HOW DIFFERENT / SO WHAT?")
r += 1
ws5.column_dimensions["A"].width = 26
ws5.column_dimensions["B"].width = 22
ws5.column_dimensions["C"].width = 95
style_col_headers(ws5, r, ["Segment", "Category", "Analysis"])
r += 1

for row in seg_df.itertuples(index=False):
    who, how, sowhat = build_analysis(row._asdict())
    sid       = row.segment_id
    fill      = FILL_SEG[sid - 1]
    seg_label = f"Segment {sid}\n{row.segment_name}"
    for cat, txt in [
        ("👤 What defines\nthis segment?",       who),
        ("📊 How do they\nbehave differently?",  how),
        ("💡 Why important\nto the business?",   sowhat),
    ]:
        for j, val in enumerate([seg_label, cat, txt], 1):
            c = ws5.cell(r, j, val)
            c.fill = fill; c.alignment = ALIGN_L
            c.border = thin_border()
            c.font = FONT_BOLD if j <= 2 else FONT_NORMAL
        ws5.row_dimensions[r].height = 65
        r += 1
    r += 1

# =========================
# SAVE
# =========================
out_dir = os.path.dirname(OUTPUT_XLSX)
if out_dir and not os.path.exists(out_dir):
    os.makedirs(out_dir, exist_ok=True)

wb.save(OUTPUT_XLSX)

print(f"\n{'='*65}")
print(f"✅ DONE  →  {OUTPUT_XLSX}")
print(f"\nSheets:")
print(f"  Task2_FICO_Threshold   — bucket + summary sentences")
print(f"  Task2_Term_Analysis    — term binary analysis")
print(f"  Task3_Segment_Hypothesis — định nghĩa trước khi nhìn data")
print(f"  Segments_Table         — quantify thực tế")
print(f"  Task5_Analysis         — written analysis")
print(f"\nFICO cut = {fico_cut} | Term = binary | DTI = descriptive")
print(f"{'='*65}")
