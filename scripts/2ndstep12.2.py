import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT  = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\week6_clean_data.csv"
OUTPUT = r"C:\Users\Tuấn Minh\OneDrive\CREDIT\risk_signals_summary.xlsx"

df = pd.read_csv(INPUT)

# =========================
# BASELINE
# =========================
baseline_high_pct = df["risk_label"].eq("high_risk").mean() * 100

# =========================
# HELPERS
# =========================
def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

FILL_HEADER = make_fill("1F4E79")
FILL_SUB    = make_fill("2E75B6")
FILL_RED    = make_fill("FF9999")   # rủi ro cao
FILL_GREEN  = make_fill("C6EFCE")   # rủi ro thấp
FILL_YELLOW = make_fill("FFEB9C")   # trung bình

FONT_WHITE  = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_ITALIC = Font(italic=True, size=10, color="595959")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, r, ncols, title):
    ws.row_dimensions[r].height = 26
    for j in range(1, ncols + 1):
        c = ws.cell(r, j)
        c.fill = FILL_HEADER; c.font = FONT_WHITE
        c.alignment = ALIGN_C; c.border = thin_border()
    ws.cell(r, 1, title)

def write_col_headers(ws, r, headers):
    ws.row_dimensions[r].height = 20
    for j, h in enumerate(headers, 1):
        c = ws.cell(r, j, h)
        c.fill = FILL_SUB; c.font = FONT_WHITE
        c.alignment = ALIGN_C; c.border = thin_border()

def write_row(ws, r, values, fill=None, bold=False):
    ws.row_dimensions[r].height = 18
    for j, v in enumerate(values, 1):
        c = ws.cell(r, j, v)
        if fill: c.fill = fill
        c.font = FONT_BOLD if bold else FONT_NORMAL
        c.alignment = ALIGN_C; c.border = thin_border()

def autofit(ws, min_w=10, max_w=50):
    for col in ws.columns:
        try:
            ltr = col[0].column_letter
        except AttributeError:
            continue
        mx = max((len(str(c.value or "")) for c in col if c.value), default=8)
        ws.column_dimensions[ltr].width = min(max(min_w, mx + 2), max_w)

# =========================
# BUILD GROUP TABLES (giữ nguyên từ code cũ)
# =========================
def build_group_table(feature_name, group_col):
    t = (
        df.groupby(group_col)
          .apply(lambda x: pd.Series({
              "n": len(x),
              "high_risk_pct": x["risk_label"].eq("high_risk").mean() * 100
          }))
          .reset_index()
          .rename(columns={group_col: "group"})
    )
    t["feature"]  = feature_name
    t["delta_pp"] = t["high_risk_pct"] - baseline_high_pct
    t["lift"]     = t["high_risk_pct"] / baseline_high_pct
    return t[["feature", "group", "n", "high_risk_pct", "delta_pp", "lift"]]

fico_tbl = build_group_table("fico_avg", "fico_group")
dti_tbl  = build_group_table("dti",      "dti_group")
term_tbl = build_group_table("term",     "term_group")
all_tbl  = pd.concat([fico_tbl, dti_tbl, term_tbl], ignore_index=True)

# Top signals
top2_high = (all_tbl.sort_values("delta_pp", ascending=False)
             .head(2)[["feature","group"]].apply(tuple, axis=1).tolist())
top2_low  = (all_tbl.sort_values("delta_pp")
             .head(2)[["feature","group"]].apply(tuple, axis=1).tolist())
top_set   = set(top2_high + top2_low)

# =========================
# TASK 1 — SIGNAL COMBINATIONS
# Câu hỏi: Kết hợp 2 tín hiệu yếu có tạo ra rủi ro mạnh hơn không?
# So sánh: 2 tín hiệu xấu vs 2 tín hiệu tốt vs 1 xấu 1 tốt
# =========================
print("Building signal combination table...")

# Dùng fico_group và dti_group (từ 2ndstep, đã có trong clean data)
combo_rows = []
for fico_g in ["low_fico", "high_fico"]:
    for dti_g in ["low_dti", "high_dti"]:
        mask  = (df["fico_group"] == fico_g) & (df["dti_group"] == dti_g)
        n     = int(mask.sum())
        if n < 100:
            continue
        pct_h = df.loc[mask, "risk_label"].eq("high_risk").mean() * 100
        combo_rows.append({
            "fico_group":   fico_g,
            "dti_group":    dti_g,
            "n":            n,
            "high_risk_pct": round(pct_h, 2),
            "delta_pp":     round(pct_h - baseline_high_pct, 2),
            "lift":         round(pct_h / baseline_high_pct, 2),
        })

combo_df = pd.DataFrame(combo_rows)

# Thêm Term vào combo (FICO × Term)
combo_term_rows = []
for fico_g in ["low_fico", "high_fico"]:
    for term_g in ["term_36", "term_60"]:
        mask  = (df["fico_group"] == fico_g) & (df["term_group"] == term_g)
        n     = int(mask.sum())
        if n < 100:
            continue
        pct_h = df.loc[mask, "risk_label"].eq("high_risk").mean() * 100
        combo_term_rows.append({
            "fico_group":   fico_g,
            "term_group":   term_g,
            "n":            n,
            "high_risk_pct": round(pct_h, 2),
            "delta_pp":     round(pct_h - baseline_high_pct, 2),
            "lift":         round(pct_h / baseline_high_pct, 2),
        })

combo_term_df = pd.DataFrame(combo_term_rows)

# Print preview
print(f"\nBaseline: {baseline_high_pct:.2f}%")
print("\nFICO × DTI combinations:")
print(combo_df.to_string(index=False))
print("\nFICO × Term combinations:")
print(combo_term_df.to_string(index=False))

# =========================
# WRITE EXCEL
# =========================
wb = Workbook()

# ── SHEET 1: Risk Signals (giữ nguyên từ code cũ, chỉ đẹp hơn) ──
ws1 = wb.active
ws1.title = "Risk_Signals"

r = 1
write_header(ws1, r, 6, f"RISK SIGNAL ANALYSIS  |  Baseline high_risk = {baseline_high_pct:.2f}%")
r += 1

for section_title, tbl in [
    ("FICO Groups", fico_tbl),
    ("DTI Groups",  dti_tbl),
    ("Term Groups", term_tbl),
]:
    r += 1
    ws1.cell(r, 1, section_title).font = FONT_BOLD
    ws1.row_dimensions[r].height = 18
    r += 1
    write_col_headers(ws1, r, ["Feature", "Group", "N", "High Risk %", "Δ vs Baseline (pp)", "Lift (x)"])
    r += 1
    for _, row in tbl.iterrows():
        key    = (row["feature"], row["group"])
        is_top = key in top_set
        delta  = row["delta_pp"]
        fill   = FILL_RED if delta > 3 else (FILL_GREEN if delta < -3 else None)
        write_row(ws1, r, [
            row["feature"], row["group"], f"{int(row['n']):,}",
            f"{row['high_risk_pct']:.2f}%", f"{row['delta_pp']:+.2f}", f"{row['lift']:.2f}x",
        ], fill=fill, bold=is_top)
        r += 1

r += 1
ws1.cell(r, 1, f"Baseline high_risk (overall): {baseline_high_pct:.2f}%").font = FONT_ITALIC
ws1.cell(r+1, 1, "In đậm = top 2 tín hiệu cao nhất và thấp nhất.").font = FONT_ITALIC
autofit(ws1)

# ── SHEET 2: TASK 1 — FICO × DTI Combinations ──
ws2 = wb.create_sheet("Task1_FICO_x_DTI")
r = 1
write_header(ws2, r, 7,
    f"TASK 1 — TỔ HỢP TÍN HIỆU: FICO × DTI  |  Baseline = {baseline_high_pct:.2f}%")
r += 1
ws2.cell(r, 1,
    "Câu hỏi: Kết hợp 2 tín hiệu xấu có tạo ra rủi ro cao hơn tổng từng tín hiệu riêng lẻ không?"
).font = FONT_ITALIC
ws2.row_dimensions[r].height = 18
r += 2

write_col_headers(ws2, r, [
    "FICO Group", "DTI Group", "N (loans)", "High Risk %",
    "Δ vs Baseline (pp)", "Lift (x)", "Diễn giải tín hiệu"
])
r += 1

signal_label = {
    ("low_fico",  "low_dti"):  "FICO xấu + DTI thấp (1 xấu, 1 tốt)",
    ("low_fico",  "mid_dti"):  "FICO xấu + DTI trung bình",
    ("low_fico",  "high_dti"): "2 tín hiệu XẤU (worst combo)",
    ("mid_fico",  "low_dti"):  "FICO tb + DTI thấp",
    ("mid_fico",  "mid_dti"):  "FICO tb + DTI trung bình (baseline)",
    ("mid_fico",  "high_dti"): "FICO tb + DTI cao",
    ("high_fico", "low_dti"):  "2 tín hiệu TỐT (best combo)",
    ("high_fico", "mid_dti"):  "FICO tốt + DTI trung bình",
    ("high_fico", "high_dti"): "FICO tốt + DTI cao (1 tốt, 1 xấu)",
}

for _, row in combo_df.sort_values(["fico_group", "dti_group"]).iterrows():
    key   = (row["fico_group"], row["dti_group"])
    label = signal_label.get(key, "")
    delta = row["delta_pp"]
    if delta > 5:      fill = FILL_RED
    elif delta < -3:   fill = FILL_GREEN
    elif abs(delta)<1: fill = FILL_YELLOW
    else:              fill = None
    bold = key in [("low_fico","high_dti"), ("high_fico","low_dti")]
    write_row(ws2, r, [
        row["fico_group"], row["dti_group"], f"{row['n']:,}",
        f"{row['high_risk_pct']:.2f}%", f"{row['delta_pp']:+.2f}",
        f"{row['lift']:.2f}x", label,
    ], fill=fill, bold=bold)
    r += 1

# Comparison highlight box
r += 1
ws2.cell(r, 1, "SO SÁNH TRỰC TIẾP:").font = FONT_BOLD
r += 1
for label, fg, dg in [
    ("2 tín hiệu XẤU (low_fico + high_dti)",  "low_fico",  "high_dti"),
    ("2 tín hiệu TỐT (high_fico + low_dti)",  "high_fico", "low_dti"),
    ("1 xấu 1 tốt   (low_fico + low_dti)",    "low_fico",  "low_dti"),
    ("1 tốt 1 xấu   (high_fico + high_dti)",  "high_fico", "high_dti"),
]:
    row_data = combo_df[(combo_df["fico_group"]==fg) & (combo_df["dti_group"]==dg)]
    if row_data.empty: continue
    rd    = row_data.iloc[0]
    delta = rd["delta_pp"]
    fill  = FILL_RED if delta > 3 else (FILL_GREEN if delta < -3 else FILL_YELLOW)
    write_row(ws2, r, [
        label, "", f"{int(rd['n']):,}",
        f"{rd['high_risk_pct']:.2f}%", f"{delta:+.2f}",
        f"{rd['lift']:.2f}x", "",
    ], fill=fill, bold=True)
    r += 1

autofit(ws2)
ws2.column_dimensions["G"].width = 40

# ── SHEET 3: TASK 1 — FICO × Term Combinations ──
ws3 = wb.create_sheet("Task1_FICO_x_Term")
r = 1
write_header(ws3, r, 7,
    f"TASK 1 — TỔ HỢP TÍN HIỆU: FICO × TERM  |  Baseline = {baseline_high_pct:.2f}%")
r += 1
ws3.cell(r, 1,
    "Câu hỏi: FICO xấu + Term dài có nguy hiểm hơn tổng 2 tín hiệu riêng lẻ không?"
).font = FONT_ITALIC
ws3.row_dimensions[r].height = 18
r += 2

write_col_headers(ws3, r, [
    "FICO Group", "Term Group", "N (loans)", "High Risk %",
    "Δ vs Baseline (pp)", "Lift (x)", "Diễn giải"
])
r += 1

signal_term_label = {
    ("high_fico", "term_36"): "2 tín hiệu TỐT (best combo)",
    ("high_fico", "term_60"): "FICO tốt + Term dài",
    ("mid_fico",  "term_36"): "FICO tb + Term ngắn",
    ("mid_fico",  "term_60"): "FICO tb + Term dài",
    ("low_fico",  "term_36"): "FICO xấu + Term ngắn",
    ("low_fico",  "term_60"): "2 tín hiệu XẤU (worst combo)",
}

for _, row in combo_term_df.sort_values(["fico_group","term_group"], ascending=[False,True]).iterrows():
    key   = (row["fico_group"], row["term_group"])
    label = signal_term_label.get(key, "")
    delta = row["delta_pp"]
    if delta > 5:    fill = FILL_RED
    elif delta < -3: fill = FILL_GREEN
    else:            fill = None
    bold = key in [("low_fico","term_60"), ("high_fico","term_36")]
    write_row(ws3, r, [
        row["fico_group"], row["term_group"], f"{row['n']:,}",
        f"{row['high_risk_pct']:.2f}%", f"{row['delta_pp']:+.2f}",
        f"{row['lift']:.2f}x", label,
    ], fill=fill, bold=bold)
    r += 1

# Comparison box
r += 1
ws3.cell(r, 1, "SO SÁNH TRỰC TIẾP:").font = FONT_BOLD
r += 1
for label, fg, tg in [
    ("2 tín hiệu XẤU (low_fico + term_60)",  "low_fico",  "term_60"),
    ("2 tín hiệu TỐT (high_fico + term_36)", "high_fico", "term_36"),
    ("1 xấu 1 tốt   (low_fico + term_36)",   "low_fico",  "term_36"),
    ("1 tốt 1 xấu   (high_fico + term_60)",  "high_fico", "term_60"),
]:
    row_data = combo_term_df[(combo_term_df["fico_group"]==fg) & (combo_term_df["term_group"]==tg)]
    if row_data.empty: continue
    rd    = row_data.iloc[0]
    delta = rd["delta_pp"]
    fill  = FILL_RED if delta > 3 else (FILL_GREEN if delta < -3 else FILL_YELLOW)
    write_row(ws3, r, [
        label, "", f"{int(rd['n']):,}",
        f"{rd['high_risk_pct']:.2f}%", f"{delta:+.2f}",
        f"{rd['lift']:.2f}x", "",
    ], fill=fill, bold=True)
    r += 1

autofit(ws3)
ws3.column_dimensions["G"].width = 40

# ── SHEET 4: TASK 1 — FICO × DTI × TERM (3-way) ──
# So sánh đúng nghĩa: 2 strong positive vs 2 strong negative signals

print("Building 3-way combination table (FICO × DTI × Term)...")

combo_3way_rows = []
for fico_g in ["low_fico", "high_fico"]:
    for dti_g in ["low_dti", "high_dti"]:
        for term_g in ["term_36", "term_60"]:
            mask = (
                (df["fico_group"] == fico_g) &
                (df["dti_group"]  == dti_g)  &
                (df["term_group"] == term_g)
            )
            n = int(mask.sum())
            if n < 100:
                continue
            pct_h = df.loc[mask, "risk_label"].eq("high_risk").mean() * 100
            combo_3way_rows.append({
                "fico_group":    fico_g,
                "dti_group":     dti_g,
                "term_group":    term_g,
                "n":             n,
                "high_risk_pct": round(pct_h, 2),
                "delta_pp":      round(pct_h - baseline_high_pct, 2),
                "lift":          round(pct_h / baseline_high_pct, 2),
            })

combo_3way_df = pd.DataFrame(combo_3way_rows)

# Định nghĩa nhãn cho từng tổ hợp
def signal_strength_label(row):
    fico  = row["fico_group"]
    dti   = row["dti_group"]
    term  = row["term_group"]

    good  = sum([fico == "high_fico", dti == "low_dti",  term == "term_36"])
    bad   = sum([fico == "low_fico",  dti == "high_dti", term == "term_60"])

    if fico == "high_fico" and dti == "low_dti"  and term == "term_36": return "3 STRONG POSITIVE (best)"
    if fico == "low_fico"  and dti == "high_dti" and term == "term_60": return "3 STRONG NEGATIVE (worst)"
    if good == 2 and bad == 0: return "2 positive + 1 neutral"
    if bad  == 2 and good == 0: return "2 negative + 1 neutral"
    if good == 2 and bad == 1: return "2 positive, 1 negative"
    if bad  == 2 and good == 1: return "2 negative, 1 positive"
    if good == 1 and bad == 1: return "mixed (1 pos, 1 neg)"
    return "neutral mix"

combo_3way_df["label"] = combo_3way_df.apply(signal_strength_label, axis=1)

ws4 = wb.create_sheet("Task1_3way_FICO_DTI_Term")
r = 1
write_header(ws4, r, 8,
    f"TASK 1 — 3-WAY: FICO × DTI × TERM  |  Baseline = {baseline_high_pct:.2f}%")
r += 1
ws4.cell(r, 1,
    "So sánh ĐÚNG: 2 strong positive (high_fico+low_dti+term_36) vs 2 strong negative (low_fico+high_dti+term_60)"
).font = FONT_ITALIC
ws4.row_dimensions[r].height = 18
r += 2

write_col_headers(ws4, r, [
    "FICO Group", "DTI Group", "Term Group",
    "N (loans)", "High Risk %", "Δ vs Baseline (pp)", "Lift (x)", "Diễn giải"
])
r += 1

# Sort: worst → best
for _, row in combo_3way_df.sort_values("high_risk_pct", ascending=False).iterrows():
    delta = row["delta_pp"]
    if   delta >= 8:   fill = make_fill("FF4D4D")   # đỏ đậm
    elif delta >= 3:   fill = FILL_RED              # đỏ nhạt
    elif delta <= -5:  fill = make_fill("70AD47")   # xanh đậm
    elif delta <= -2:  fill = FILL_GREEN            # xanh nhạt
    else:              fill = FILL_YELLOW

    is_extreme = row["label"] in [
        "3 STRONG POSITIVE (best)",
        "3 STRONG NEGATIVE (worst)"
    ]
    ws4.row_dimensions[r].height = 18
    for j, v in enumerate([
        row["fico_group"], row["dti_group"], row["term_group"],
        f"{int(row['n']):,}", f"{row['high_risk_pct']:.2f}%",
        f"{row['delta_pp']:+.2f}", f"{row['lift']:.2f}x", row["label"],
    ], 1):
        c = ws4.cell(r, j, v)
        c.fill = fill
        c.font = FONT_BOLD if is_extreme else FONT_NORMAL
        c.alignment = ALIGN_C; c.border = thin_border()
    r += 1

# ── Comparison box: 4 tổ hợp quan trọng nhất ──
r += 1
ws4.cell(r, 1, "SO SÁNH TRỰC TIẾP — 4 TỔ HỢP QUAN TRỌNG NHẤT:").font = FONT_BOLD
r += 1
write_col_headers(ws4, r, [
    "Mô tả", "", "",
    "N (loans)", "High Risk %", "Δ vs Baseline (pp)", "Lift (x)", ""
])
r += 1

key_combos = [
    ("3 STRONG NEGATIVE", "low_fico",  "high_dti", "term_60"),
    ("2 neg + 1 neutral (low_fico+high_dti+term_36)", "low_fico",  "high_dti", "term_36"),
    ("2 neg + 1 neutral (low_fico+low_dti+term_60)",  "low_fico",  "low_dti",  "term_60"),
    ("3 STRONG POSITIVE", "high_fico", "low_dti",  "term_36"),
]

for label, fg, dg, tg in key_combos:
    rd = combo_3way_df[
        (combo_3way_df["fico_group"] == fg) &
        (combo_3way_df["dti_group"]  == dg) &
        (combo_3way_df["term_group"] == tg)
    ]
    if rd.empty: continue
    rd   = rd.iloc[0]
    delta = rd["delta_pp"]
    fill  = make_fill("FF4D4D") if delta >= 8 else (make_fill("70AD47") if delta <= -5 else FILL_YELLOW)
    ws4.row_dimensions[r].height = 18
    for j, v in enumerate([
        label, "", "",
        f"{int(rd['n']):,}", f"{rd['high_risk_pct']:.2f}%",
        f"{rd['delta_pp']:+.2f}", f"{rd['lift']:.2f}x", "",
    ], 1):
        c = ws4.cell(r, j, v)
        c.fill = fill; c.font = FONT_BOLD
        c.alignment = ALIGN_C; c.border = thin_border()
    r += 1

# Autofit
for col in ws4.columns:
    try:
        ltr = col[0].column_letter
    except AttributeError:
        continue
    mx = max((len(str(c.value or "")) for c in col if c.value), default=8)
    ws4.column_dimensions[ltr].width = min(max(10, mx + 2), 55)
ws4.column_dimensions["H"].width = 45

# =========================
# SAVE
# =========================
wb.save(OUTPUT)
print(f"\nBaseline high-risk %: {round(baseline_high_pct, 4)}")
print(f"Saved to: {OUTPUT}")
print(f"\nSheets: Risk_Signals | Task1_FICO_x_DTI | Task1_FICO_x_Term | Task1_3way_FICO_DTI_Term")
